"""Excel export builders — Final Report (7 sheets) and DMRB Report (12 sheets).

Styling matches the export_reference implementation (src/io_excel.py, Reports copy/clean_up.py):

  DMRB Report  — each table at start_col=2, start_row=5 (first table);
                 2-row merged title above each table (bold 20pt, thick border);
                 white bold headers; all data cells centered; column-A spacer (width 4).

  Final Report — tables at A1; white bold headers; centered; no title rows above tables.

Reference palette (src/io_excel.py):
  GREEN  C6EFCE · PINK  FFC7CE · NOTICE BDD7EE · GRAY  D9D9D9 · YELLOW FFFFCC
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from typing import Any, Callable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from domain import turnover_lifecycle as tl
from services.exports.export_service import is_finite_numeric

# ── Palette (reference: src/io_excel.py) ──────────────────────────────────────

_C_GREEN  = "C6EFCE"
_C_PINK   = "FFC7CE"
_C_NOTICE = "BDD7EE"
_C_GRAY   = "D9D9D9"
_C_YELLOW = "FFFFCC"

_DATE_FMT  = "MM/DD/YY"
_SLA_DAYS  = 10          # calendar-day SLA threshold

# En-dash bucket labels to match reference output exactly
_AGING_BUCKETS       = ("1\u201310", "11\u201320", "21\u201330", "31\u201360", "61\u2013120", "120+")
_SHORT_AGING_BUCKETS = ("1\u201310", "11\u201320")   # Active Aging only shows first 2

_VACANT_STATES = {tl.PHASE_VACANT_NOT_READY, tl.PHASE_VACANT_READY}
_NOTICE_STATES = {tl.PHASE_ON_NOTICE, tl.PHASE_PRE_NOTICE}

_STATE_TO_STATUS: dict[str, str] = {
    tl.PHASE_VACANT_READY:     "Vacant Ready",
    tl.PHASE_VACANT_NOT_READY: "Vacant Not Ready",
    tl.PHASE_ON_NOTICE:        "On Notice",
    tl.PHASE_PRE_NOTICE:       "Pre-Notice",
    tl.PHASE_OCCUPIED:         "Occupied",
}

_TASK_LABEL: dict[str, str] = {
    "INSPECT":         "Inspection",
    "CARPET_BID":      "Carpet Bid",
    "MAKE_READY_BID":  "Make Ready Bid",
    "PAINT":           "Paint",
    "MAKE_READY":      "Make Ready",
    "HOUSEKEEPING":    "Housekeeping",
    "CARPET_CLEAN":    "Carpet Clean",
    "FINAL_WALK":      "Final Walk",
    "QUALITY_CONTROL": "Quality Control",
}

_TASK_STATUS_READABLE: dict[str, str] = {
    "COMPLETED":   "Done",
    "IN_PROGRESS": "In Progress",
    "SCHEDULED":   "Scheduled",
    "NOT_STARTED": "Not Started",
}

# (row_key, section_title, sched_date_col, vendor_date_col, status_col)
_SCHED_TASKS = [
    ("task_insp",         "Inspections",   "task_insp_sched",   "task_insp_vendor",   "task_insp_status"),
    ("task_paint",        "Paint",         "task_paint_sched",  "task_paint_vendor",  "task_paint_status"),
    ("task_mr",           "Make Ready",    "task_mr_sched",     "task_mr_vendor",     "task_mr_status"),
    ("task_hk",           "House Keeping", "task_hk_sched",     "task_hk_vendor",     "task_hk_status"),
    ("task_carpet_clean", "Carpet Clean",  "task_cc_sched",     "task_cc_vendor",     "task_cc_status"),
]

# ── Low-level openpyxl helpers ────────────────────────────────────────────────

def _apply_fill(cell, hex_color: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)


_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium15",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
_WHITE_BOLD   = Font(bold=True, color="FFFFFF")
_CENTER       = Alignment(horizontal="center", vertical="center")
_THICK_SIDE   = Side(style="thick")
_THICK_BORDER = Border(
    top=_THICK_SIDE, bottom=_THICK_SIDE,
    left=_THICK_SIDE, right=_THICK_SIDE,
)

# ── Fill helpers (reference rules) ────────────────────────────────────────────

def _status_fill_color(status: str | None) -> str | None:
    if not status:
        return None
    s = str(status).strip().lower()
    if s.startswith("on notice") or s == "pre-notice":
        return _C_NOTICE
    if s == "vacant ready":
        return _C_GREEN
    if s == "vacant not ready":
        return _C_PINK
    return None


def _apply_status_fills(
    ws,
    start_row: int,
    end_row: int,
    col_map: dict[str, int],
) -> None:
    """Post-write fill pass: mirrors reference _apply_dashboard_formatting."""
    status_col = col_map.get("Status")
    dv_col     = col_map.get("Days Vacant") or col_map.get("DV")
    alert_col  = col_map.get("\u26a0\ufe0f Alert")   # ⚠️ Alert
    mi_col     = col_map.get("M-I Date") or col_map.get("M-I")

    if status_col is None:
        return

    for row_idx in range(start_row + 1, end_row + 1):
        raw = ws.cell(row=row_idx, column=status_col).value
        if not raw:
            continue
        color = _status_fill_color(str(raw))
        if color:
            _apply_fill(ws.cell(row=row_idx, column=status_col), color)

        s = str(raw).strip().lower()
        is_not_ready = (s == "vacant not ready")
        is_ready     = (s == "vacant ready")
        is_notice    = s.startswith("on notice") or s == "pre-notice"

        if dv_col:
            dv_val = ws.cell(row=row_idx, column=dv_col).value
            if dv_val is not None:
                if is_not_ready:
                    try:
                        dv_compliant = float(dv_val) < _SLA_DAYS
                    except (TypeError, ValueError):
                        dv_compliant = True
                    _apply_fill(
                        ws.cell(row=row_idx, column=dv_col),
                        _C_GREEN if dv_compliant else _C_PINK,
                    )
                elif color:
                    _apply_fill(ws.cell(row=row_idx, column=dv_col), color)

        if alert_col:
            av = ws.cell(row=row_idx, column=alert_col).value
            if av is not None and str(av).strip():
                av_lower = str(av).strip().lower()
                if "work stalled" in av_lower:
                    _apply_fill(ws.cell(row=row_idx, column=alert_col), _C_GRAY)
                elif "needs attention" in av_lower:
                    _apply_fill(ws.cell(row=row_idx, column=alert_col), _C_YELLOW)
                elif color:
                    _apply_fill(ws.cell(row=row_idx, column=alert_col), color)

        if mi_col:
            mv = ws.cell(row=row_idx, column=mi_col).value
            if mv is not None and color:
                _apply_fill(ws.cell(row=row_idx, column=mi_col), color)


def _apply_aging_fills(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Parse 'Unit X (Status) | DV-N' cells and apply reference fills."""
    for row_idx in range(start_row + 1, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not val:
                continue
            text = str(val).strip().lower()
            m = re.search(r'\(([^)]+)\)', text)
            if not m:
                continue
            status = m.group(1).strip()
            dv_m   = re.search(r'dv-(\d+)', text)
            dv_num = int(dv_m.group(1)) if dv_m else 0
            if status.startswith("on notice"):
                _apply_fill(ws.cell(row=row_idx, column=col_idx), _C_NOTICE)
            elif status == "vacant ready":
                _apply_fill(ws.cell(row=row_idx, column=col_idx), _C_GREEN)
            elif status == "vacant not ready":
                _apply_fill(
                    ws.cell(row=row_idx, column=col_idx),
                    _C_PINK if dv_num >= _SLA_DAYS else _C_GREEN,
                )


# ── Scalar / date helpers ─────────────────────────────────────────────────────

def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _safe_val(v: Any) -> Any:
    """Convert numpy/pandas scalars to Python natives; NaN/NaT → None."""
    if v is None:
        return None
    try:
        import numpy as np
        if isinstance(v, float) and v != v:
            return None
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            return float(v)
        if isinstance(v, (np.bool_,)):
            return bool(v)
    except ImportError:
        pass
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _aging_bucket(vd: int) -> str:
    if vd <= 10:
        return "1\u201310"
    if vd <= 20:
        return "11\u201320"
    if vd <= 30:
        return "21\u201330"
    if vd <= 60:
        return "31\u201360"
    if vd <= 120:
        return "61\u2013120"
    return "120+"


def _task_exec_status(task: Any) -> str:
    if not isinstance(task, dict):
        return ""
    raw = task.get("execution_status") or ""
    return _TASK_STATUS_READABLE.get(raw, raw)


def _task_sched_date(task: Any) -> date | None:
    if not isinstance(task, dict):
        return None
    return _parse_date(task.get("scheduled_date"))


def _task_vendor_date(task: Any) -> date | None:
    if not isinstance(task, dict):
        return None
    return _parse_date(task.get("vendor_due_date"))


def _task_label_str(task: Any) -> str:
    if not isinstance(task, dict):
        return "—"
    tt = task.get("task_type") or ""
    return _TASK_LABEL.get(tt, tt) or "—"


def _nvm(row: dict) -> str:
    if row.get("move_in_date") is not None:
        return "M"
    state = row.get("operational_state", "")
    if state in _VACANT_STATES:
        return "V"
    if state in _NOTICE_STATES:
        return "N"
    return ""


def _current_task_date(row: dict) -> date | None:
    ct = row.get("current_task")
    if not isinstance(ct, dict):
        return None
    task_type = ct.get("task_type", "")
    key_map = {
        "INSPECT":       "task_insp",
        "PAINT":         "task_paint",
        "MAKE_READY":    "task_mr",
        "HOUSEKEEPING":  "task_hk",
        "CARPET_CLEAN":  "task_carpet_clean",
        "FINAL_WALK":    "task_fw",
    }
    key = key_map.get(task_type)
    if key:
        return _task_sched_date(row.get(key))
    return None


def _next_task_date(row: dict) -> date | None:
    nt = row.get("next_task")
    if not isinstance(nt, dict):
        return None
    task_type = nt.get("task_type", "")
    key_map = {
        "INSPECT":       "task_insp",
        "PAINT":         "task_paint",
        "MAKE_READY":    "task_mr",
        "HOUSEKEEPING":  "task_hk",
        "CARPET_CLEAN":  "task_carpet_clean",
        "FINAL_WALK":    "task_fw",
    }
    key = key_map.get(task_type)
    if key:
        return _task_sched_date(row.get(key))
    return None


# ── DataFrame builder ─────────────────────────────────────────────────────────

def _rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Convert export rows to a DataFrame with all derived columns pre-computed."""
    _extra = [
        "status", "mo_confirm", "dv_bucket",
        "is_vacant", "is_notice", "nvm",
        "current_task_label", "next_task_label",
        "current_task_date", "next_task_date",
        "task_state", "in_turn_execution", "has_notes",
        "task_insp_status",  "task_paint_status",  "task_mr_status",
        "task_hk_status",    "task_cc_status",
        "task_insp_sched",   "task_paint_sched",   "task_mr_sched",
        "task_hk_sched",     "task_cc_sched",
        "task_insp_vendor",  "task_paint_vendor",  "task_mr_vendor",
        "task_hk_vendor",    "task_cc_vendor",
    ]
    if not rows:
        base_cols = [
            "turnover_id", "unit_code", "phase", "dv", "dtbr", "days_to_move_in",
            "operational_state", "move_out_date", "move_in_date", "report_ready_date",
            "available_date", "confirmed_move_out_date", "scheduled_move_out_date",
            "attention_badge", "sla_breach", "plan_breach", "inspection_sla_breach",
            "sla_movein_breach", "is_task_stalled", "task_completion_ratio", "is_unit_ready",
            "wd_summary", "notes_joined", "qc_status",
            "task_insp", "task_paint", "task_mr", "task_hk", "task_carpet_clean",
            "task_fw", "task_qc", "task_carpet_bid", "task_mr_bid",
            "current_task", "next_task", "tasks_list",
        ]
        return pd.DataFrame(columns=base_cols + _extra)

    df = pd.DataFrame(rows)

    # Status label
    df["status"] = df["operational_state"].map(_STATE_TO_STATUS).fillna(df["operational_state"])

    # MO/Confirm
    df["mo_confirm"] = df["confirmed_move_out_date"].apply(
        lambda v: "Yes" if v is not None else ""
    )

    # DV bucket
    df["dv_bucket"] = df["dv"].apply(
        lambda v: _aging_bucket(int(v)) if is_finite_numeric(v) else ""
    )

    # Boolean state helpers
    df["is_vacant"] = df["operational_state"].isin(_VACANT_STATES)
    df["is_notice"] = df["operational_state"].isin(_NOTICE_STATES)

    # N/V/M code
    df["nvm"] = df.apply(_nvm, axis=1)

    # Task labels
    df["current_task_label"] = df["current_task"].apply(_task_label_str)
    df["next_task_label"]    = df["next_task"].apply(_task_label_str)

    # Task dates (current & next, from their scheduled_date)
    df["current_task_date"] = df.apply(_current_task_date, axis=1)
    df["next_task_date"]    = df.apply(_next_task_date, axis=1)

    # Task state
    def _task_state(r: Any) -> str:
        ratio = r["task_completion_ratio"] or 0.0
        if ratio >= 100:
            return "All Tasks Complete"
        if ratio > 0:
            return "In Progress"
        return "Not Started"
    df["task_state"] = df.apply(_task_state, axis=1)

    # In-turn execution flag (vacant + actively in progress)
    df["in_turn_execution"] = df["is_vacant"] & (df["task_completion_ratio"].fillna(0) > 0)

    # Has notes
    df["has_notes"] = df["notes_joined"].fillna("").astype(bool)

    # Per-task status and date columns
    for key, _title, sched_col, vendor_col, status_col in _SCHED_TASKS:
        df[status_col] = df[key].apply(_task_exec_status)
        df[sched_col]  = df[key].apply(_task_sched_date)
        df[vendor_col] = df[key].apply(_task_vendor_date)

    return df


# ── Column selector ───────────────────────────────────────────────────────────

def _pick(df: pd.DataFrame, col_map: dict[str, str]) -> pd.DataFrame:
    """Select and rename columns; missing source cols become empty strings."""
    result: dict[str, Any] = {}
    for src, dst in col_map.items():
        if src in df.columns:
            result[dst] = df[src].values
        else:
            result[dst] = [""] * len(df)
    return pd.DataFrame(result)


# ── DMRB write layer ─────────────────────────────────────────────────────────

def _is_date_col(series: pd.Series) -> bool:
    for v in series.dropna():
        return isinstance(v, (date, datetime))
    return False


def _write_entry(
    ws,
    df: pd.DataFrame,
    start_col: int,
    start_row: int,
    table_name: str,
    title: str | None = None,
    fills_fn: Callable | None = None,
) -> int:
    """Write df as a titled formatted table starting at (start_col, start_row).

    The 2 rows above start_row are used for the merged section title.
    Returns the next available start_row for the following entry.
    """
    end_col = start_col + len(df.columns) - 1
    end_row = start_row + len(df)

    # ── Title (2 rows merged above start_row)
    if title:
        merge_top = start_row - 2
        merge_bot = start_row - 1
        ws.merge_cells(
            start_row=merge_top, start_column=start_col,
            end_row=merge_bot,   end_column=end_col,
        )
        title_cell = ws.cell(row=merge_top, column=start_col, value=title)
        title_cell.font      = Font(bold=True, size=20)
        title_cell.alignment = _CENTER
        for r in range(merge_top, merge_bot + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = _THICK_BORDER

    # ── Detect date columns
    date_col_idx: set[int] = set()
    for ci, col_name in enumerate(df.columns):
        if _is_date_col(df[col_name]):
            date_col_idx.add(ci)

    # ── Header row
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell            = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font       = _WHITE_BOLD
        cell.alignment  = _CENTER

    # ── Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ci    = col_idx - start_col
            value = _safe_val(value)
            if value is None:
                cell = ws.cell(row=row_idx, column=col_idx, value=None)
            elif ci in date_col_idx:
                if isinstance(value, date) and not isinstance(value, datetime):
                    value = datetime(value.year, value.month, value.day)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = _DATE_FMT
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER

    # ── Excel table (only register when there is at least 1 data row;
    #    header-only tables — ref spans a single row — are invalid in Excel
    #    and trigger the "We found a problem" repair dialog on open)
    if len(df) > 0:
        ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        tbl               = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = _TABLE_STYLE
        ws.add_table(tbl)

    # ── Optional post-write fills
    if fills_fn and len(df) > 0:
        col_map = {col_name: start_col + ci for ci, col_name in enumerate(df.columns)}
        fills_fn(ws, start_row, end_row, col_map)

    # next start_row = +len(df) data rows + 1 blank gap + 4 (title block)
    return start_row + len(df) + 1 + 4


def _autofit_dmrb(ws, start_col: int, max_col: int) -> None:
    """Column A → width 4 (spacer); table columns → content width + 3."""
    for col_num in range(1, max_col + 1):
        col_letter = get_column_letter(col_num)
        if col_num < start_col:
            ws.column_dimensions[col_letter].width = 4.0
        else:
            max_len = 0
            for row in ws.iter_rows(min_col=col_num, max_col=col_num, values_only=True):
                if row[0] is not None:
                    max_len = max(max_len, len(str(row[0])))
            ws.column_dimensions[col_letter].width = max_len + 3


# ── Final Report write layer ──────────────────────────────────────────────────

def _write_fr_table(
    ws,
    df: pd.DataFrame,
    table_name: str,
    start_row: int = 1,
    start_col: int = 1,
) -> None:
    """Write a Final-Report-style table (A1-based, white bold headers, centered, no title)."""
    end_col = start_col + len(df.columns) - 1
    end_row = start_row + len(df)

    date_col_idx: set[int] = set()
    for ci, col_name in enumerate(df.columns):
        if _is_date_col(df[col_name]):
            date_col_idx.add(ci)

    # Header
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell           = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font      = _WHITE_BOLD
        cell.alignment = _CENTER

    # Data
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ci    = col_idx - start_col
            value = _safe_val(value)
            if value is None:
                cell = ws.cell(row=row_idx, column=col_idx, value=None)
            elif ci in date_col_idx:
                if isinstance(value, date) and not isinstance(value, datetime):
                    value = datetime(value.year, value.month, value.day)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = _DATE_FMT
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER

    # Excel table (only when there is at least 1 data row — header-only tables are invalid)
    if len(df) > 0:
        ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        tbl               = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = _TABLE_STYLE
        ws.add_table(tbl)


def _autofit_fr(ws, df: pd.DataFrame) -> None:
    """Final Report autofit: max(header_len, data_len) + 2."""
    for i, col in enumerate(df.columns, start=1):
        max_data = df[col].fillna("").astype(str).str.len().max() if len(df) > 0 else 0
        ws.column_dimensions[get_column_letter(i)].width = max(max_data, len(str(col))) + 2


# ── Aging bucket grid helper ──────────────────────────────────────────────────

def _build_bucket_df(
    rows: list[dict],
    buckets: tuple[str, ...] = _AGING_BUCKETS,
) -> pd.DataFrame:
    """Build a bucket-grid DataFrame matching reference build_aging_tables format."""
    from collections import defaultdict
    columns: dict[str, list[str]] = {b: [] for b in buckets}
    for r in rows:
        dv = r.get("dv")
        if not is_finite_numeric(dv):
            continue
        b = _aging_bucket(int(dv))
        if b not in columns:
            continue
        state = r.get("operational_state", "")
        status_label = _STATE_TO_STATUS.get(state, state)
        entry = f"{r['unit_code']} ({status_label}) | DV-{int(dv)}"
        columns[b].append(entry)

    # Sort each bucket by DV descending
    for b in buckets:
        columns[b].sort(
            key=lambda e: -int(e.split("DV-")[1]) if "DV-" in e else 0
        )

    max_len = max((len(v) for v in columns.values()), default=0)
    for k in columns:
        columns[k] += [""] * (max_len - len(columns[k]))

    return pd.DataFrame(columns) if max_len > 0 else pd.DataFrame({b: [] for b in buckets})


# ══════════════════════════════════════════════════════════════════════════════
# BUILD FINAL REPORT (7 sheets)
# ══════════════════════════════════════════════════════════════════════════════

def build_final_report(rows: list[dict]) -> bytes:
    """Build the 7-sheet Final Report workbook from export rows."""
    df = _rows_to_df(rows)
    buf = io.BytesIO()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Reconciliation
    recon = _pick(df, {
        "phase":               "Phase",
        "unit_code":           "Unit",
        "status":              "Status",
        "available_date":      "Available Date",
        "report_ready_date":   "Move-In Ready Date",
        "move_in_date":        "Move In Date",
        "mo_confirm":          "MO / Confirm",
    })
    ws1 = wb.create_sheet("Reconciliation")
    _write_fr_table(ws1, recon, "Reconciliation")
    _autofit_fr(ws1, recon)

    # ── Sheet 2: Split View
    ws2 = wb.create_sheet("Split View")
    has_mi  = df[df["move_in_date"].notna()] if not df.empty else df
    no_mi   = df[df["move_in_date"].isna()]  if not df.empty else df
    sv_cols = {
        "phase": "Phase", "unit_code": "Unit", "status": "Status",
        "available_date": "Available Date", "report_ready_date": "Move-In Ready Date",
        "move_in_date": "Move In Date", "mo_confirm": "MO / Confirm",
    }
    title_font = Font(bold=True, size=13, color="1F4E79")
    n_sv_cols  = len(sv_cols)

    # "Has Move In" section
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_sv_cols)
    cell = ws2.cell(row=1, column=1, value="Has Move In")
    cell.font = title_font
    sv_mi = _pick(has_mi, sv_cols)
    _write_fr_table(ws2, sv_mi, "SplitHasMI", start_row=2)

    # "No Move In" section — gap row then section
    gap_row    = 3 + len(sv_mi)
    title_row  = gap_row + 1
    data_start = title_row + 1
    ws2.merge_cells(
        start_row=title_row, start_column=1,
        end_row=title_row, end_column=n_sv_cols,
    )
    cell = ws2.cell(row=title_row, column=1, value="No Move In")
    cell.font = title_font
    sv_no = _pick(no_mi, sv_cols)
    _write_fr_table(ws2, sv_no, "SplitNoMI", start_row=data_start)
    _autofit_fr(ws2, sv_mi)

    # ── Sheet 3: Available Units
    vacant = df[df["is_vacant"]] if not df.empty else df
    avail  = _pick(vacant, {
        "phase":             "Phase",
        "unit_code":         "Unit",
        "status":            "Status",
        "available_date":    "Available Date",
        "report_ready_date": "Move-In Ready Date",
    })
    ws3 = wb.create_sheet("Available Units")
    _write_fr_table(ws3, avail, "AvailableUnits")
    _autofit_fr(ws3, avail)

    # ── Sheet 4: Move Ins
    has_movein = df[df["move_in_date"].notna()] if not df.empty else df
    move_ins   = _pick(has_movein, {
        "phase":        "Phase",
        "unit_code":    "Unit",
        "move_in_date": "Move In Date",
    })
    ws4 = wb.create_sheet("Move Ins")
    _write_fr_table(ws4, move_ins, "MoveIns")
    _autofit_fr(ws4, move_ins)

    # ── Sheet 5: Move Outs
    move_outs = _pick(df, {
        "phase":          "Phase",
        "unit_code":      "Unit",
        "available_date": "Move-Out Date",
    })
    ws5 = wb.create_sheet("Move Outs")
    _write_fr_table(ws5, move_outs, "MoveOuts")
    _autofit_fr(ws5, move_outs)

    # ── Sheet 6: Pending FAS
    notice  = df[df["is_notice"]] if not df.empty else df
    pend_df = _pick(notice, {
        "phase":                   "Phase",
        "unit_code":               "Unit",
        "available_date":          "MO / Cancel Date",
        "scheduled_move_out_date": "Lease End",
    })
    pend_df["Completed"] = ""
    ws6 = wb.create_sheet("Pending FAS")
    _write_fr_table(ws6, pend_df, "PendingFAS")
    _autofit_fr(ws6, pend_df)

    # ── Sheet 7: Move Activity
    activity = _pick(df, {
        "phase":          "Phase",
        "unit_code":      "Unit",
        "available_date": "Move-Out Date",
        "move_in_date":   "Move In Date",
    })
    ws7 = wb.create_sheet("Move Activity")
    _write_fr_table(ws7, activity, "MoveActivity")
    _autofit_fr(ws7, activity)

    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# BUILD DMRB REPORT (12 sheets)
# ══════════════════════════════════════════════════════════════════════════════

_SC = 2    # start_col for all DMRB entries
_SR = 5    # start_row for first entry on each sheet


def build_dmrb_report(rows: list[dict], metrics: dict, today: date) -> bytes:
    """Build the 12-sheet DMRB Report workbook from export rows."""
    df = _rows_to_df(rows)
    buf = io.BytesIO()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Dashboard
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard")
    dash = _pick(df, {
        "phase":          "Phase",
        "unit_code":      "Unit",
        "status":         "Status",
        "dv":             "Days Vacant",
        "move_in_date":   "M-I Date",
        "attention_badge": "\u26a0\ufe0f Alert",
    })
    end_row = _SR + len(dash)
    _write_entry(ws, dash, _SC, _SR, "DashboardVacant",
                 title="Units Overview",
                 fills_fn=_apply_status_fills)
    _autofit_dmrb(ws, _SC, _SC + len(dash.columns) - 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Aging (3 bucket-grid sections)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Aging")
    cur_row = _SR

    status_norm = df["status"].astype(str).str.strip().str.upper() if not df.empty else pd.Series([], dtype=str)
    all_rows     = rows
    ready_rows   = [r for r, s in zip(rows, status_norm) if s == "VACANT READY"]    if not df.empty else []
    not_rdy_rows = [r for r, s in zip(rows, status_norm) if s == "VACANT NOT READY"] if not df.empty else []

    for title, rlist, tname in [
        ("Aging Buckets \u2014 All Units",          all_rows,     "AgingAll"),
        ("Aging Buckets \u2014 Vacant Ready",        ready_rows,   "AgingReady"),
        ("Aging Buckets \u2014 Vacant Not Ready",    not_rdy_rows, "AgingNotReady"),
    ]:
        grid = _build_bucket_df(rlist)
        next_row = _write_entry(ws, grid, _SC, cur_row, tname, title=title)
        # apply aging fills
        if len(grid) > 0:
            _apply_aging_fills(ws, cur_row, cur_row + len(grid), _SC, _SC + len(grid.columns) - 1)
        cur_row = next_row

    _autofit_dmrb(ws, _SC, _SC + len(_AGING_BUCKETS) - 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 3 — Active Aging (short buckets, one section per phase)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Active Aging")
    cur_row = _SR

    vnr_rows = [r for r in rows if r.get("operational_state") == tl.PHASE_VACANT_NOT_READY]
    phases   = sorted({r["phase"] for r in vnr_rows}) if vnr_rows else []
    for phase_name in phases:
        phase_rows = [r for r in vnr_rows if r["phase"] == phase_name]
        grid = _build_bucket_df(phase_rows, _SHORT_AGING_BUCKETS)
        _phase_num = re.sub(r"[^0-9A-Za-z]", "", str(phase_name))
        tname = f"ActiveAgingP{_phase_num}"
        next_row = _write_entry(
            ws, grid, _SC, cur_row, tname,
            title=f"Active Aging \u2014 {phase_name} (Vacant Not Ready)",
        )
        if len(grid) > 0:
            _apply_aging_fills(ws, cur_row, cur_row + len(grid), _SC, _SC + len(grid.columns) - 1)
        cur_row = next_row

    if not phases:
        empty_grid = pd.DataFrame({b: [] for b in _SHORT_AGING_BUCKETS})
        _write_entry(ws, empty_grid, _SC, _SR, "ActiveAgingEmpty",
                     title="Active Aging (Vacant Not Ready)")

    _autofit_dmrb(ws, _SC, _SC + len(_SHORT_AGING_BUCKETS) - 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 4 — Operations (4 sections: matching reference run.py)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Operations")
    cur_row = _SR

    # 1) Move-In Dashboard: has move-in AND not unit ready
    mi_df = df[df["move_in_date"].notna() & ~df["is_unit_ready"].astype(bool)] if not df.empty else df
    mi_df = mi_df.sort_values("move_in_date", ascending=True, na_position="last") if not mi_df.empty else mi_df
    mi_view = _pick(mi_df, {
        "phase":               "Phase",
        "unit_code":           "Unit",
        "status":              "Status",
        "available_date":      "M-O",
        "dv":                  "DV",
        "move_in_date":        "M-I",
        "dtbr":                "DTBR",
        "nvm":                 "N/V/M",
        "task_completion_ratio": "Progress %",
        "current_task_label":  "Current Task",
        "status":              "Execution Reason",
        "attention_badge":     "\u26a0\ufe0f Alert",
    })
    # Fix: status appears twice; map operational_state as "Execution Reason"
    mi_view2 = pd.DataFrame({
        "Phase":            mi_df["phase"].values                   if not mi_df.empty else [],
        "Unit":             mi_df["unit_code"].values               if not mi_df.empty else [],
        "Status":           mi_df["status"].values                  if not mi_df.empty else [],
        "M-O":              mi_df["available_date"].values          if not mi_df.empty else [],
        "DV":               mi_df["dv"].values                      if not mi_df.empty else [],
        "M-I":              mi_df["move_in_date"].values            if not mi_df.empty else [],
        "DTBR":             mi_df["dtbr"].values                    if not mi_df.empty else [],
        "N/V/M":            mi_df["nvm"].values                     if not mi_df.empty else [],
        "Progress %":       mi_df["task_completion_ratio"].values   if not mi_df.empty else [],
        "Current Task":     mi_df["current_task_label"].values      if not mi_df.empty else [],
        "Execution Reason": mi_df["operational_state"].values       if not mi_df.empty else [],
        "\u26a0\ufe0f Alert": mi_df["attention_badge"].values       if not mi_df.empty else [],
    })
    cur_row = _write_entry(ws, mi_view2, _SC, cur_row, "OpsMoveIn",
                           title="Move In Dashboard",
                           fills_fn=_apply_status_fills)

    # 2) SLA Breach
    sla_df = df[df["sla_breach"].astype(bool)] if not df.empty else df
    sla_df = sla_df.sort_values("dv", ascending=False, na_position="last") if not sla_df.empty else sla_df
    sla_view = _pick(sla_df, {
        "phase":               "Phase",
        "unit_code":           "Unit",
        "status":              "Status",
        "available_date":      "M-O",
        "dv":                  "DV",
        "nvm":                 "N/V/M",
        "current_task_label":  "Current Task",
        "task_completion_ratio": "Progress %",
        "attention_badge":     "\u26a0\ufe0f Alert",
        "notes_joined":        "Notes Category",
    })
    cur_row = _write_entry(ws, sla_view, _SC, cur_row, "OpsSLABreach",
                           title="SLA Breach (Turn Overdue Over 10 Days)",
                           fills_fn=_apply_status_fills)

    # 3) Inspection SLA Breach
    insp_df = df[df["inspection_sla_breach"].astype(bool)] if not df.empty else df
    insp_df = insp_df.sort_values("dv", ascending=False, na_position="last") if not insp_df.empty else insp_df
    insp_view = _pick(insp_df, {
        "phase":              "Phase",
        "unit_code":          "Unit",
        "status":             "Status",
        "available_date":     "M-O",
        "task_insp_sched":    "Insp Date",
        "task_insp_status":   "Inspection",
        "attention_badge":    "\u26a0\ufe0f Alert",
    })
    cur_row = _write_entry(ws, insp_view, _SC, cur_row, "OpsInspBreach",
                           title="Inspection SLA Breach",
                           fills_fn=_apply_status_fills)

    # 4) Plan Breach
    plan_df = df[df["plan_breach"].astype(bool)] if not df.empty else df
    plan_df = plan_df.sort_values("dv", ascending=False, na_position="last") if not plan_df.empty else plan_df
    plan_view = _pick(plan_df, {
        "phase":              "Phase",
        "unit_code":          "Unit",
        "report_ready_date":  "Ready Date",
        "status":             "Status",
        "status":             "Execution State",
        "current_task_label": "Execution Reason",
        "attention_badge":    "\u26a0\ufe0f Alert",
    })
    # Plan Breach: Phase, Unit, Ready Date, Status, Execution State, Execution Reason, ⚠️ Alert
    plan_view2 = pd.DataFrame({
        "Phase":             plan_df["phase"].values               if not plan_df.empty else [],
        "Unit":              plan_df["unit_code"].values            if not plan_df.empty else [],
        "Ready Date":        plan_df["report_ready_date"].values    if not plan_df.empty else [],
        "Status":            plan_df["status"].values               if not plan_df.empty else [],
        "Execution State":   plan_df["operational_state"].values    if not plan_df.empty else [],
        "Execution Reason":  plan_df["current_task_label"].values   if not plan_df.empty else [],
        "\u26a0\ufe0f Alert": plan_df["attention_badge"].values     if not plan_df.empty else [],
    })
    cur_row = _write_entry(ws, plan_view2, _SC, cur_row, "OpsPlanBreach",
                           title="Plan Breach",
                           fills_fn=_apply_status_fills)

    _autofit_dmrb(ws, _SC, _SC + 11)  # generous max col estimate

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 5 — Walking Path Board (vacant only)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Walking Path Board")
    vacant_df = df[df["is_vacant"]].sort_values("unit_code") if not df.empty else df
    wpb = _pick(vacant_df, {
        "phase":              "Phase",
        "phase":              "Building",   # no separate B column; use phase
        "unit_code":          "Unit",
        "status":             "Status",
        "attention_badge":    "Move Ins \u26a0\ufe0f Alert",
        "task_insp_status":   "Inspection",
        "task_paint_status":  "Paint",
        "task_mr_status":     "MR",
        "task_hk_status":     "HK",
        "task_cc_status":     "CC",
        "qc_status":          "QC",
    })
    # Rebuild properly (avoid duplicate Phase key issue)
    wpb_data: dict[str, Any] = {
        "Phase":                    vacant_df["phase"].values    if not vacant_df.empty else [],
        "Building":                 vacant_df["phase"].values    if not vacant_df.empty else [],
        "Unit":                     vacant_df["unit_code"].values if not vacant_df.empty else [],
        "Status":                   vacant_df["status"].values   if not vacant_df.empty else [],
        "Move Ins \u26a0\ufe0f Alert": vacant_df["attention_badge"].values if not vacant_df.empty else [],
        "Inspection":               vacant_df["task_insp_status"].values   if not vacant_df.empty else [],
        "Paint":                    vacant_df["task_paint_status"].values   if not vacant_df.empty else [],
        "MR":                       vacant_df["task_mr_status"].values      if not vacant_df.empty else [],
        "HK":                       vacant_df["task_hk_status"].values      if not vacant_df.empty else [],
        "CC":                       vacant_df["task_cc_status"].values      if not vacant_df.empty else [],
        "QC":                       vacant_df["qc_status"].values           if not vacant_df.empty else [],
        "Notes":                    [""] * (len(vacant_df) if not vacant_df.empty else 0),
    }
    wpb_df = pd.DataFrame(wpb_data)
    _write_entry(ws, wpb_df, _SC, _SR, "WalkBoard",
                 title="Walking Path Board",
                 fills_fn=_apply_status_fills)
    _autofit_dmrb(ws, _SC, _SC + len(wpb_df.columns) - 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 6 — Tasks (3 sections: Stalled, Clean Turn, Task Flow)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Tasks")
    cur_row = _SR

    # 1) Task Stalled
    stalled_df = df[df["is_task_stalled"].astype(bool)] if not df.empty else df
    stalled_df = stalled_df.sort_values("task_completion_ratio", ascending=True, na_position="last") if not stalled_df.empty else stalled_df
    stalled_view = pd.DataFrame({
        "Phase":          stalled_df["phase"].values          if not stalled_df.empty else [],
        "Unit":           stalled_df["unit_code"].values      if not stalled_df.empty else [],
        "Current Task":   stalled_df["current_task_label"].values if not stalled_df.empty else [],
        "Progress %":     stalled_df["task_completion_ratio"].values if not stalled_df.empty else [],
        "Execution State": stalled_df["status"].values        if not stalled_df.empty else [],
        "\u26a0\ufe0f Alert": stalled_df["attention_badge"].values if not stalled_df.empty else [],
    })
    cur_row = _write_entry(ws, stalled_view, _SC, cur_row, "TaskStalled",
                           title="Task Stalled",
                           fills_fn=_apply_status_fills)

    # 2) Clean Turn: task_state == "All Tasks Complete" AND vacant (reference: Task_State == "All Tasks Complete")
    clean_df = df[(df["task_state"] == "All Tasks Complete") & df["is_vacant"]] if not df.empty else df
    clean_df = clean_df.sort_values("unit_code") if not clean_df.empty else clean_df
    clean_view = pd.DataFrame({
        "Phase":           clean_df["phase"].values           if not clean_df.empty else [],
        "Unit":            clean_df["unit_code"].values       if not clean_df.empty else [],
        "Status":          clean_df["status"].values          if not clean_df.empty else [],
        "Execution State": clean_df["task_state"].values      if not clean_df.empty else [],
        "Progress %":      clean_df["task_completion_ratio"].values if not clean_df.empty else [],
    })
    cur_row = _write_entry(ws, clean_view, _SC, cur_row, "CleanTurn",
                           title="Clean Turn",
                           fills_fn=_apply_status_fills)

    # 3) Task Flow: in-turn-execution AND in progress
    flow_df = df[df["in_turn_execution"] & (df["task_state"] == "In Progress")] if not df.empty else df
    flow_df = flow_df.sort_values("task_completion_ratio", ascending=True, na_position="last") if not flow_df.empty else flow_df
    flow_view = pd.DataFrame({
        "Phase":      flow_df["phase"].values             if not flow_df.empty else [],
        "Unit":       flow_df["unit_code"].values         if not flow_df.empty else [],
        "Task":       flow_df["current_task_label"].values if not flow_df.empty else [],
        "Task Date":  flow_df["current_task_date"].values  if not flow_df.empty else [],
        "Next":       flow_df["next_task_label"].values    if not flow_df.empty else [],
        "Next Date":  flow_df["next_task_date"].values     if not flow_df.empty else [],
        "Progress %": flow_df["task_completion_ratio"].values if not flow_df.empty else [],
    })
    cur_row = _write_entry(ws, flow_view, _SC, cur_row, "TaskFlow", title="Task Flow")
    _autofit_dmrb(ws, _SC, _SC + 6)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 7 — Schedule (per task type: Phase, Unit, Date, Status)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Schedule")
    cur_row = _SR
    max_sched_col = _SC + 3

    for _key, title, sched_col, _vendor_col, status_col in _SCHED_TASKS:
        if not df.empty:
            status_norm_s = df[status_col].fillna("").str.strip().str.lower()
            sec = df[
                df[sched_col].notna() &
                status_norm_s.isin({"in progress", "scheduled", "done", "completed", ""}).map(
                    lambda x: not x  # keep rows where status is NOT empty & date exists
                )
            ].copy() if False else df[df[sched_col].notna()].copy()
            # Filter: scheduled_date not null AND status is active (not done)
            sec = df[
                df[sched_col].notna() &
                df[status_col].fillna("").str.strip().str.lower().isin(
                    {"in progress", "scheduled", "in_progress"}
                )
            ].copy()
            sec = sec.sort_values(sched_col, ascending=True, na_position="last")
        else:
            sec = df

        sched_data: dict = {
            "Phase":  sec["phase"].values      if not sec.empty else [],
            "Unit":   sec["unit_code"].values  if not sec.empty else [],
            "Date":   sec[sched_col].values    if not sec.empty else [],
            "Status": sec[status_col].values   if not sec.empty else [],
        }
        if title == "Make Ready":
            sched_data["Assigned To"] = [""] * (len(sec) if not sec.empty else 0)
        sched_view = pd.DataFrame(sched_data)
        tname = f"Sched{title.replace(' ', '')}"
        cur_row = _write_entry(ws, sched_view, _SC, cur_row, tname, title=title)
        max_sched_col = max(max_sched_col, _SC + len(sched_view.columns) - 1)

    _autofit_dmrb(ws, _SC, max_sched_col)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 8 — Upcoming (vendor_due_date within 0-7 days, per task type)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Upcoming")
    cur_row = _SR
    window_end = today + timedelta(days=7)
    max_up_col = _SC + 3

    for _key, title, _sched_col, vendor_col, status_col in _SCHED_TASKS:
        if not df.empty:
            sec = df[
                df[vendor_col].apply(
                    lambda d: d is not None and today <= d <= window_end
                )
            ].copy()
            sec = sec.sort_values(vendor_col, ascending=True, na_position="last")
        else:
            sec = df

        up_data: dict = {
            "Phase":  sec["phase"].values      if not sec.empty else [],
            "Unit":   sec["unit_code"].values  if not sec.empty else [],
            "Date":   sec[vendor_col].values   if not sec.empty else [],
            "Status": sec[status_col].values   if not sec.empty else [],
        }
        if title == "Make Ready":
            up_data["Assigned To"] = [""] * (len(sec) if not sec.empty else 0)
        up_view = pd.DataFrame(up_data)
        tname = f"Up{title.replace(' ', '')}"
        cur_row = _write_entry(ws, up_view, _SC, cur_row, tname, title=title)
        max_up_col = max(max_up_col, _SC + len(up_view.columns) - 1)

    _autofit_dmrb(ws, _SC, max_up_col)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 9 — WD Audit (vacant only)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("WD Audit")
    vacant_s = df[df["is_vacant"]].sort_values("unit_code") if not df.empty else df
    wd_df = pd.DataFrame({
        "Phase": vacant_s["phase"].values    if not vacant_s.empty else [],
        "Unit":  vacant_s["unit_code"].values if not vacant_s.empty else [],
        "W/D":   vacant_s["wd_summary"].values if not vacant_s.empty else [],
    })
    _write_entry(ws, wd_df, _SC, _SR, "WDAudit", title="W/D Audit")
    _autofit_dmrb(ws, _SC, _SC + 2)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 10 — Daily Ops (7 sections matching reference daily_ops.py)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Daily Ops")
    cur_row = _SR

    total_u  = len(df)
    vacant_n = int(df["is_vacant"].sum()) if not df.empty else 0
    vac_rate = (vacant_n / total_u * 100) if total_u > 0 else 0.0

    # 1) Portfolio Overview
    po_df = pd.DataFrame({
        "Metric": ["Total Units", "Vacant Units", "Occupied Units", "Vacancy Rate"],
        "Value":  [total_u, vacant_n, total_u - vacant_n, f"{vac_rate:.1f}%"],
    })
    title_str = f"Portfolio Overview \u2014 {today.strftime('%A, %B %d, %Y')}"
    cur_row = _write_entry(ws, po_df, _SC, cur_row, "PortfolioOverview", title=title_str)

    # 2) Turn Performance
    vacant_dv = df.loc[df["is_vacant"], "dv"].dropna() if not df.empty else pd.Series([], dtype=float)
    avg_turn    = float(vacant_dv.mean())    if len(vacant_dv) > 0 else 0.0
    median_turn = float(vacant_dv.median()) if len(vacant_dv) > 0 else 0.0
    max_turn    = float(vacant_dv.max())    if len(vacant_dv) > 0 else 0.0
    tp_df = pd.DataFrame({
        "Metric": ["Average Turn Time", "Median Turn Time", "Longest Turn"],
        "Value":  [
            f"{avg_turn:.1f} days",
            f"{median_turn:.1f} days",
            f"{max_turn:.0f} days" if max_turn else "\u2014",
        ],
    })
    cur_row = _write_entry(ws, tp_df, _SC, cur_row, "TurnPerformance", title="Turn Performance")

    # 3) SLA Compliance
    tot_breach   = int(df["sla_breach"].sum())            if not df.empty else 0
    insp_breach  = int(df["inspection_sla_breach"].sum()) if not df.empty else 0
    mi_breach    = int(df["sla_movein_breach"].sum())     if not df.empty else 0
    plan_breach  = int(df["plan_breach"].sum())           if not df.empty else 0
    comp_rate    = (
        (vacant_n - tot_breach) / vacant_n * 100 if vacant_n > 0 else 100.0
    )
    sla_df = pd.DataFrame({
        "Metric": ["Turn SLA Breaches", "Inspection SLA Breaches",
                   "Move-In Risk", "Plan Breaches", "Compliance Rate"],
        "Value":  [tot_breach, insp_breach, mi_breach, plan_breach, f"{comp_rate:.1f}%"],
    })
    cur_row = _write_entry(ws, sla_df, _SC, cur_row, "SLACompliance", title="SLA Compliance")

    # 4) Operational Status (use operational_state mapped to label, matching reference)
    if not df.empty:
        state_counts = df["operational_state"].map(_STATE_TO_STATUS).fillna(df["operational_state"]).value_counts()
        op_metrics = list(state_counts.index)
        op_values  = [
            f"{cnt} ({cnt / total_u * 100:.1f}%)" for cnt in state_counts.values
        ]
    else:
        op_metrics, op_values = [], []
    ops_status_df = pd.DataFrame({"State": op_metrics, "Units": op_values})
    cur_row = _write_entry(ws, ops_status_df, _SC, cur_row, "OperationalStatus",
                           title="Operational Status")

    # 5) Work In Progress
    in_progress = df[df["in_turn_execution"]] if not df.empty else df
    avg_comp    = float(in_progress["task_completion_ratio"].mean()) if not in_progress.empty else 0.0
    stalled_n   = int(df["is_task_stalled"].sum()) if not df.empty else 0
    wip_df = pd.DataFrame({
        "Metric": ["Units in Active Turn", "Avg Completion", "Stalled Tasks"],
        "Value":  [len(in_progress), f"{avg_comp:.1f}%", stalled_n],
    })
    cur_row = _write_entry(ws, wip_df, _SC, cur_row, "WorkInProgress", title="Work In Progress")

    # 6) Ready For Move-In
    ready_n   = int(df["is_unit_ready"].sum()) if not df.empty else 0
    ready_df2 = pd.DataFrame({
        "Metric": ["Units Ready", "Units QC Complete", "Awaiting QC"],
        "Value":  [ready_n, ready_n, 0],   # schema-blocked: no separate QC-complete flag
    })
    cur_row = _write_entry(ws, ready_df2, _SC, cur_row, "ReadyForMoveIn",
                           title="Ready For Move-In")

    # 7) Critical Alerts
    at_risk_n = int(
        (df["is_vacant"] & ~df["is_unit_ready"].astype(bool) &
         (df["dv"].fillna(0) >= 8) & (df["dv"].fillna(0) < 10)).sum()
    ) if not df.empty else 0
    alerts: list[tuple[str, str]] = []
    if at_risk_n  > 0: alerts.append((f"{at_risk_n} units at risk of SLA breach (8-9 days)", "Warning"))
    if tot_breach > 0: alerts.append((f"{tot_breach} units currently in SLA breach", "Critical"))
    if mi_breach  > 0: alerts.append((f"{mi_breach} units at risk of missing move-in", "Critical"))
    if stalled_n  > 0: alerts.append((f"{stalled_n} units with stalled tasks", "Warning"))
    if not alerts:     alerts.append(("No critical alerts \u2014 operations running smoothly", "OK"))
    alerts_df = pd.DataFrame({"Alert": [a[0] for a in alerts], "Severity": [a[1] for a in alerts]})
    cur_row = _write_entry(ws, alerts_df, _SC, cur_row, "CriticalAlerts", title="Critical Alerts")

    _autofit_dmrb(ws, _SC, _SC + 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 11 — Priority (2 tables: scoring legend + top units)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Priority")
    cur_row = _SR

    # Table 1: Scoring legend
    score_df = pd.DataFrame({
        "Condition": ["SLA Breach", "Move-In Risk", "Task Stalled", "Aging >= 8 days", "Has Notes/Issues"],
        "Points":    [50, 40, 30, 20, 10],
    })
    cur_row = _write_entry(ws, score_df, _SC, cur_row, "PriorityScoring", title="Priority Scoring")

    # Table 2: Top priority units (score-based, vacant only)
    if not df.empty:
        pri = df[df["is_vacant"]].copy()
        pri["_score"] = 0
        pri.loc[pri["sla_breach"].astype(bool),       "_score"] += 50
        pri.loc[pri["sla_movein_breach"].astype(bool), "_score"] += 40
        pri.loc[pri["is_task_stalled"].astype(bool),   "_score"] += 30
        pri.loc[(pri["dv"].fillna(0) >= 8),            "_score"] += 20
        pri.loc[pri["has_notes"],                      "_score"] += 10
        top = pri.nlargest(20, "_score").reset_index(drop=True)

        def _build_flags(r: Any) -> str:
            f: list[str] = []
            if r["sla_breach"]:       f.append("SLA BREACH")
            if r["sla_movein_breach"]: f.append("MOVE-IN RISK")
            if r["is_task_stalled"]:  f.append("STALLED")
            if r["has_notes"]:        f.append(str(r["notes_joined"])[:30])
            return " | ".join(f) if f else ""

        def _build_action(r: Any) -> str:
            if r["sla_breach"]:
                return "ESCALATE - Already in breach, needs immediate resolution"
            if r["sla_movein_breach"]:
                dtm = r["days_to_move_in"]
                ds  = f"{int(dtm)}" if is_finite_numeric(dtm) else "?"
                return f"URGENT - Move-in {ds} days away, unit not ready"
            if r["is_task_stalled"]:
                return "REVIEW - Task stalled, investigate blockers"
            return "MONITOR - Track progress closely"

        top["Alerts"] = top.apply(_build_flags,  axis=1)
        top["Action"] = top.apply(_build_action, axis=1)
        pri_view = pd.DataFrame({
            "Phase":        top["phase"].values,
            "Priority":     top["_score"].values,
            "Days Vacant":  top["dv"].values,
            "Status":       top["status"].values,
            "Progress %":   top["task_completion_ratio"].values,
            "Current Task": top["current_task_label"].values,
            "Alerts":       top["Alerts"].values,
            "Action":       top["Action"].values,
        })
    else:
        pri_view = pd.DataFrame(columns=[
            "Phase", "Priority", "Days Vacant", "Status",
            "Progress %", "Current Task", "Alerts", "Action",
        ])

    n_top  = len(pri_view)
    cur_row = _write_entry(ws, pri_view, _SC, cur_row, "TopPriorityUnits",
                           title=f"Top {n_top} Priority Units")
    _autofit_dmrb(ws, _SC, _SC + max(len(score_df.columns), len(pri_view.columns)) - 1)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 12 — Phase Performance (3 tables matching reference phase_performance.py)
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Phase Performance")
    cur_row = _SR

    if not df.empty:
        perf = df.groupby("phase", sort=True).agg(
            Units=("unit_code",             "count"),
            Vacant=("is_vacant",            "sum"),
            Avg_Turn=("dv",                 "mean"),
            SLA_Breaches=("sla_breach",     "sum"),
            Ready=("is_unit_ready",         "sum"),
            Avg_Completion=("task_completion_ratio", "mean"),
        ).round(1).reset_index()
        perf.rename(columns={"phase": "Phase"}, inplace=True)
        perf["Vacancy Rate"]    = (perf["Vacant"] / perf["Units"] * 100).round(1)
        perf["Compliance Rate"] = (
            (perf["Vacant"] - perf["SLA_Breaches"]) / perf["Vacant"].replace(0, 1) * 100
        ).round(1)
        def _indicator(avg: Any) -> str:
            if avg is None or (isinstance(avg, float) and avg != avg):
                return ""
            return "Good" if avg <= 6 else "Watch" if avg <= 8 else "Critical"
        perf["Performance"] = perf["Avg_Turn"].apply(_indicator)
        perf = perf.sort_values("Avg_Turn", na_position="last")
        comp = perf[[
            "Phase", "Units", "Vacant", "Avg_Turn", "SLA_Breaches",
            "Ready", "Avg_Completion", "Vacancy Rate", "Compliance Rate", "Performance",
        ]].rename(columns={
            "Avg_Turn":       "Avg Turn Days",
            "SLA_Breaches":   "SLA Breaches",
            "Avg_Completion": "Avg Completion %",
        })
    else:
        comp = pd.DataFrame(columns=[
            "Phase", "Units", "Vacant", "Avg Turn Days", "SLA Breaches",
            "Ready", "Avg Completion %", "Vacancy Rate", "Compliance Rate", "Performance",
        ])

    cur_row = _write_entry(ws, comp, _SC, cur_row, "PhaseComparison",
                           title="Phase Performance Comparison")

    # Portfolio totals
    if not comp.empty:
        totals = pd.DataFrame({
            "Metric": [
                "Total Units", "Total Vacant", "Portfolio Avg Turn",
                "Total SLA Breaches", "Total Ready", "Portfolio Avg Completion",
                "Portfolio Compliance Rate",
            ],
            "Value": [
                len(df),
                int(comp["Vacant"].sum()),
                f"{comp['Avg Turn Days'].mean():.1f} days",
                int(comp["SLA Breaches"].sum()),
                int(comp["Ready"].sum()),
                f"{comp['Avg Completion %'].mean():.1f}%",
                f"{comp['Compliance Rate'].mean():.1f}%",
            ],
        })
    else:
        totals = pd.DataFrame({"Metric": [], "Value": []})

    cur_row = _write_entry(ws, totals, _SC, cur_row, "PortfolioSummary",
                           title="Portfolio Summary")

    # Best & Worst performers
    if len(comp) >= 2:
        best  = comp.head(2)
        worst = comp.tail(2).iloc[::-1]
        perf_rows: list[dict] = []
        for _, r in best.iterrows():
            perf_rows.append({
                "Phase": r["Phase"], "Category": "Best Performer",
                "Avg Turn Days": r["Avg Turn Days"],
                "Compliance Rate": f"{r['Compliance Rate']}%",
            })
        for _, r in worst.iterrows():
            perf_rows.append({
                "Phase": r["Phase"], "Category": "Needs Attention",
                "Avg Turn Days": r["Avg Turn Days"],
                "Compliance Rate": f"{r['Compliance Rate']}%",
            })
        performers = pd.DataFrame(perf_rows)
    else:
        performers = pd.DataFrame(
            columns=["Phase", "Category", "Avg Turn Days", "Compliance Rate"]
        )

    cur_row = _write_entry(ws, performers, _SC, cur_row, "Performers",
                           title="Best & Worst Performers")

    _autofit_dmrb(ws, _SC, _SC + max(len(comp.columns), 4) - 1)

    wb.save(buf)
    return buf.getvalue()
