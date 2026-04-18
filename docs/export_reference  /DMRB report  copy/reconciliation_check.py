import os
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----------------------------------------
# PATHS (relative to this script's location)
# ----------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DMRB_PATH = os.path.join(SCRIPT_DIR, "data", "DMRB_raw.xlsx")
RECON_PATH = os.path.join(SCRIPT_DIR, "output", "Final_Report.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "output", "reconciliation_output.xlsx")

# ----------------------------------------
# LOAD DATA
# ----------------------------------------
# Reconciliation (source of truth)
recon = pd.read_excel(RECON_PATH, sheet_name="Reconciliation")
recon["Unit"] = recon["Unit"].str.strip()
recon["Available Date"] = pd.to_datetime(recon["Available Date"], errors="coerce")
recon["Move In Date"] = pd.to_datetime(recon["Move In Date"], errors="coerce")

# DMRB sheet
dmrb = pd.read_excel(DMRB_PATH, sheet_name="DMRB ")
dmrb.columns = dmrb.columns.str.strip()
dmrb["Unit"] = dmrb["Unit"].str.strip()
dmrb["Move_out"] = pd.to_datetime(dmrb["Move_out"], errors="coerce")
dmrb["Move_in"] = pd.to_datetime(dmrb["Move_in"], errors="coerce")
dmrb["Ready_Date"] = pd.to_datetime(dmrb["Ready_Date"], errors="coerce")

# ----------------------------------------
# BUILD COMPARISON
# ----------------------------------------
# Start from Reconciliation as the source of truth
df = recon[["phase", "Unit", "Status", "Available Date", "Move-In Ready Date",
            "Move In Date", "MO / Confirm"]].copy()

# Merge DMRB data
dmrb_subset = dmrb[["Unit", "Move_out", "Ready_Date", "Move_in"]].rename(columns={
    "Move_out": "DMRB Available Date",
    "Ready_Date": "DMRB Ready Date",
    "Move_in": "DMRB Move In Date",
})
df = df.merge(dmrb_subset, on="Unit", how="left")

# ----------------------------------------
# FLAG MISMATCHES
# ----------------------------------------
# Unit missing from DMRB entirely
df["Missing in DMRB"] = df["DMRB Available Date"].isna() & df["DMRB Move In Date"].isna()

# Available Date mismatch (ignore differences <= 2 days)
df["Avail Date Diff"] = (df["Available Date"] - df["DMRB Available Date"]).abs()
df["Avail Date Mismatch"] = (
    df["Available Date"].notna()
    & df["DMRB Available Date"].notna()
    & (df["Avail Date Diff"] > pd.Timedelta(days=2))
)

# Move In Date mismatch (ignore differences <= 2 days)
df["Move In Diff"] = (df["Move In Date"] - df["DMRB Move In Date"]).abs()
df["Move In Mismatch"] = (
    df["Move In Date"].notna()
    & df["DMRB Move In Date"].notna()
    & (df["Move In Diff"] > pd.Timedelta(days=2))
)

# Move In present in Recon but missing in DMRB (unit exists but date is blank)
df["Move In Missing in DMRB"] = (
    df["Move In Date"].notna()
    & ~df["Missing in DMRB"]
    & df["DMRB Move In Date"].isna()
)

# Available Date present in Recon but missing in DMRB
df["Avail Date Missing in DMRB"] = (
    df["Available Date"].notna()
    & ~df["Missing in DMRB"]
    & df["DMRB Available Date"].isna()
)

# Any issue at all
df["Has Issue"] = (
    df["Missing in DMRB"]
    | df["Avail Date Mismatch"]
    | df["Move In Mismatch"]
    | df["Move In Missing in DMRB"]
    | df["Avail Date Missing in DMRB"]
)

# Build a human-readable issue column
def describe_issue(row):
    issues = []
    if row["Missing in DMRB"]:
        issues.append("Unit not in DMRB")
    if row["Avail Date Mismatch"]:
        issues.append(
            f"Avail Date: Recon={row['Available Date'].strftime('%m/%d/%y')} "
            f"vs DMRB={row['DMRB Available Date'].strftime('%m/%d/%y')}"
        )
    if row["Avail Date Missing in DMRB"]:
        issues.append("Avail Date in Recon but blank in DMRB")
    if row["Move In Mismatch"]:
        issues.append(
            f"Move In: Recon={row['Move In Date'].strftime('%m/%d/%y')} "
            f"vs DMRB={row['DMRB Move In Date'].strftime('%m/%d/%y')}"
        )
    if row["Move In Missing in DMRB"]:
        issues.append("Move In in Recon but blank in DMRB")
    return " | ".join(issues)

df["Issue"] = df.apply(describe_issue, axis=1)

# ----------------------------------------
# OUTPUT – only rows with issues
# ----------------------------------------
issues_df = df[df["Has Issue"]].copy()

# Select output columns: same as Reconciliation + Issue
out_cols = [
    "phase", "Unit", "Status",
    "Available Date", "Move-In Ready Date",
    "Move In Date", "MO / Confirm", "Issue",
]
issues_df = issues_df[out_cols].sort_values(
    ["Move In Date", "Available Date"], ascending=[True, True]
).reset_index(drop=True)

# ----------------------------------------
# WRITE STYLED EXCEL
# ----------------------------------------
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    issues_df.to_excel(writer, sheet_name="Mismatches", index=False)

    # ----------------------------------------
    # SHEET 2 – Units not in DMRB
    # ----------------------------------------
    not_in_dmrb = df[df["Missing in DMRB"]].copy()
    not_in_dmrb = not_in_dmrb.rename(columns={
        "Available Date": "Move Out",
        "Move-In Ready Date": "Ready Date",
        "Move In Date": "Move In",
    })
    status_map = {
        "Vacant Not Ready": "Vacant not ready",
        "Vacant ready": "Vacant ready",
        "Vacant not ready": "Vacant not ready",
        "On notice": "On Notice ",
        "On notice (Break)": "On Notice ",
    }
    not_in_dmrb["Status"] = not_in_dmrb["Status"].map(status_map).fillna(not_in_dmrb["Status"])
    not_in_dmrb["DV"] = ""
    not_in_dmrb = not_in_dmrb[["Unit", "Status", "Move Out", "Ready Date", "DV", "Move In"]].sort_values(
        ["Move Out", "Move In"], ascending=[True, True]
    ).reset_index(drop=True)
    not_in_dmrb.to_excel(writer, sheet_name="Not in DMRB", index=False)

    ws2 = writer.sheets["Not in DMRB"]

    for i, col in enumerate(not_in_dmrb.columns, start=1):
        max_len = max(
            not_in_dmrb[col].fillna("").astype(str).str.len().max(), len(col)
        ) + 2
        ws2.column_dimensions[get_column_letter(i)].width = min(max_len, 30)

    for col_idx, col in enumerate(not_in_dmrb.columns, start=1):
        if col in ("Move Out", "Ready Date", "Move In"):
            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "MM/DD/YY"

    last_row2 = len(not_in_dmrb) + 1
    last_col2 = get_column_letter(len(not_in_dmrb.columns))
    table2 = Table(displayName="NotInDMRB", ref=f"A1:{last_col2}{last_row2}")
    table2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws2.add_table(table2)

    center2 = Alignment(horizontal="center", vertical="center")
    for row in ws2.iter_rows(min_row=1, max_row=last_row2, min_col=1, max_col=len(not_in_dmrb.columns)):
        for cell in row:
            cell.alignment = center2

    header_font2 = Font(color="FFFFFF", bold=True)
    for cell in ws2[1]:
        cell.font = header_font2

    wb = writer.book
    ws = writer.sheets["Mismatches"]

    # Autofit columns
    for i, col in enumerate(issues_df.columns, start=1):
        max_len = max(
            issues_df[col].fillna("").astype(str).str.len().max(), len(col)
        ) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 45)

    # Date formatting
    for col_idx, col in enumerate(issues_df.columns, start=1):
        if "Date" in col:
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "MM/DD/YY"

    # Excel table
    last_row = len(issues_df) + 1
    last_col = get_column_letter(len(issues_df.columns))
    table = Table(displayName="Mismatches", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Center alignment
    center = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=len(issues_df.columns)):
        for cell in row:
            cell.alignment = center

    # Header styling
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Highlight the Issue column with a light red fill
    issue_col = list(issues_df.columns).index("Issue") + 1
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=issue_col, max_col=issue_col):
        for cell in row:
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

print(f"Total Reconciliation units: {len(recon)}")
print(f"Total DMRB units: {len(dmrb)}")
print(f"Units with issues: {len(issues_df)}")
print(f"\nOutput saved to: {OUTPUT_PATH}")

if len(issues_df) > 0:
    print("\nIssue summary:")
    print(f"  Missing in DMRB entirely: {df['Missing in DMRB'].sum()}")
    print(f"  Available Date mismatch:  {df['Avail Date Mismatch'].sum()}")
    print(f"  Available Date missing:   {df['Avail Date Missing in DMRB'].sum()}")
    print(f"  Move In Date mismatch:    {df['Move In Mismatch'].sum()}")
    print(f"  Move In Date missing:     {df['Move In Missing in DMRB'].sum()}")
else:
    print("\n✅ Everything matches!")
