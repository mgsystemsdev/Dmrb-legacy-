import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def read_excel_table(path, sheet_name, table_name):
    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        target = sheet_name.strip()
        matches = [s for s in wb.sheetnames if s.strip() == target]
        if len(matches) == 1:
            sheet_name = matches[0]
        else:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

    ws = wb[sheet_name]

    if table_name not in ws.tables:
        target = table_name.strip()
        matches = [t for t in ws.tables if t.strip() == target]
        if len(matches) == 1:
            table_name = matches[0]
        else:
            raise ValueError(
                f"Table '{table_name}' not found in sheet '{sheet_name}'. "
                f"Available: {list(ws.tables.keys())}"
            )

    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    rows = ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True
    )

    rows = list(rows)
    headers = rows[0]
    data = rows[1:]

    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip() for c in df.columns]

    dupes = df.columns[df.columns.duplicated(keep=False)]
    if len(dupes):
        raise ValueError(f"Duplicate columns after stripping whitespace: {dupes.tolist()}")

    return df


# ============================================================
# CONDITIONAL FORMATTING — Dashboard
# ============================================================

SLA_DAYS = 10

FILL_GREEN = None
FILL_PINK = None
FILL_NOTICE = None
FILL_GRAY = None
FILL_YELLOW = None


def _init_fills():
    from openpyxl.styles import PatternFill
    global FILL_GREEN, FILL_PINK, FILL_NOTICE, FILL_GRAY, FILL_YELLOW
    FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_PINK = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_NOTICE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")


def _apply_dashboard_formatting(ws, start_row, end_row, col_map):
    _init_fills()
    today = datetime.today()

    status_col = col_map.get("Status")
    dv_col = col_map.get("Days Vacant")
    mi_col = col_map.get("M-I Date")
    alert_col = col_map.get("⚠️ Alert")

    if status_col is None:
        return

    for row_idx in range(start_row + 1, end_row + 1):
        raw_status = ws.cell(row=row_idx, column=status_col).value
        if raw_status is None or str(raw_status).strip() == "":
            continue

        status_lower = str(raw_status).strip().lower()
        is_notice = status_lower.startswith("on notice")
        is_ready = status_lower == "vacant ready"
        is_not_ready = status_lower == "vacant not ready"

        alert_text = ""
        if alert_col:
            av = ws.cell(row=row_idx, column=alert_col).value
            if av is not None:
                alert_text = str(av).strip().lower()
        is_stalled = "work stalled" in alert_text
        is_needs_attention = "needs attention" in alert_text

        # --- Status column ---
        if is_notice:
            ws.cell(row=row_idx, column=status_col).fill = FILL_NOTICE
        elif is_ready:
            ws.cell(row=row_idx, column=status_col).fill = FILL_GREEN
        elif is_not_ready:
            ws.cell(row=row_idx, column=status_col).fill = FILL_PINK

        # --- Shared SLA check for Vacant Not Ready ---
        dv_compliant = True
        if is_not_ready and dv_col:
            dv_raw = ws.cell(row=row_idx, column=dv_col).value
            try:
                dv_compliant = float(dv_raw) < SLA_DAYS
            except (TypeError, ValueError):
                dv_compliant = True

        # --- Days Vacant column ---
        if dv_col:
            dv_val = ws.cell(row=row_idx, column=dv_col).value
            if dv_val is None:
                pass
            elif is_notice:
                ws.cell(row=row_idx, column=dv_col).fill = FILL_NOTICE
            elif is_ready:
                ws.cell(row=row_idx, column=dv_col).fill = FILL_GREEN
            elif is_not_ready:
                ws.cell(row=row_idx, column=dv_col).fill = FILL_PINK if not dv_compliant else FILL_GREEN

        # --- Alert column (gray=stalled, yellow=needs attention, else mirrors DV) ---
        if alert_col:
            alert_val = ws.cell(row=row_idx, column=alert_col).value
            if alert_val is not None and str(alert_val).strip() != "":
                if is_notice:
                    ws.cell(row=row_idx, column=alert_col).fill = FILL_NOTICE
                elif is_ready:
                    ws.cell(row=row_idx, column=alert_col).fill = FILL_GREEN
                elif is_not_ready:
                    if not dv_compliant:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_PINK
                    elif is_stalled:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_GRAY
                    elif is_needs_attention:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_YELLOW
                    else:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_GREEN

        # --- M-I Date column (uses Days Vacant for SLA check) ---
        if mi_col:
            mi_val = ws.cell(row=row_idx, column=mi_col).value
            if mi_val is None:
                pass
            elif is_notice:
                ws.cell(row=row_idx, column=mi_col).fill = FILL_NOTICE
            elif is_ready:
                ws.cell(row=row_idx, column=mi_col).fill = FILL_GREEN
            elif is_not_ready:
                ws.cell(row=row_idx, column=mi_col).fill = FILL_PINK if not dv_compliant else FILL_GREEN


def _apply_operations_formatting(ws, start_row, end_row, col_map):
    """Apply same conditional formatting rules to Operations tables."""
    _init_fills()

    status_col = col_map.get("Status")
    dv_col = col_map.get("DV")
    alert_col = col_map.get("⚠️ Alert")

    if status_col is None:
        return

    for row_idx in range(start_row + 1, end_row + 1):
        raw_status = ws.cell(row=row_idx, column=status_col).value
        if raw_status is None or str(raw_status).strip() == "":
            continue

        status_lower = str(raw_status).strip().lower()
        is_notice = status_lower.startswith("on notice")
        is_ready = status_lower == "vacant ready"
        is_not_ready = status_lower == "vacant not ready"

        alert_text = ""
        if alert_col:
            av = ws.cell(row=row_idx, column=alert_col).value
            if av is not None:
                alert_text = str(av).strip().lower()
        is_stalled = "work stalled" in alert_text
        is_needs_attention = "needs attention" in alert_text

        # --- Status column ---
        if is_notice:
            ws.cell(row=row_idx, column=status_col).fill = FILL_NOTICE
        elif is_ready:
            ws.cell(row=row_idx, column=status_col).fill = FILL_GREEN
        elif is_not_ready:
            ws.cell(row=row_idx, column=status_col).fill = FILL_PINK

        # --- DV column ---
        dv_compliant = True
        if dv_col:
            dv_val = ws.cell(row=row_idx, column=dv_col).value
            if dv_val is not None:
                try:
                    dv_compliant = float(dv_val) < SLA_DAYS
                except (TypeError, ValueError):
                    dv_compliant = True
                if is_notice:
                    ws.cell(row=row_idx, column=dv_col).fill = FILL_NOTICE
                elif is_ready:
                    ws.cell(row=row_idx, column=dv_col).fill = FILL_GREEN
                elif is_not_ready:
                    ws.cell(row=row_idx, column=dv_col).fill = FILL_PINK if not dv_compliant else FILL_GREEN

        # --- Alert column ---
        if alert_col:
            alert_val = ws.cell(row=row_idx, column=alert_col).value
            if alert_val is not None and str(alert_val).strip() != "":
                if is_notice:
                    ws.cell(row=row_idx, column=alert_col).fill = FILL_NOTICE
                elif is_ready:
                    ws.cell(row=row_idx, column=alert_col).fill = FILL_GREEN
                elif is_not_ready:
                    if not dv_compliant:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_PINK
                    elif is_stalled:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_GRAY
                    elif is_needs_attention:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_YELLOW
                    else:
                        ws.cell(row=row_idx, column=alert_col).fill = FILL_GREEN


def _apply_aging_formatting(ws, start_row, end_row, start_col, end_col):
    """Apply conditional formatting to Aging sheet cells.

    Each cell contains: "Unit xxx (Status) | DV-N"
    Parse status and DV to apply the same color rules as Dashboard.
    """
    import re
    _init_fills()

    for row_idx in range(start_row + 1, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None or str(val).strip() == "":
                continue

            text = str(val).strip().lower()

            # Extract status from parentheses
            m = re.search(r'\(([^)]+)\)', text)
            if not m:
                continue
            status = m.group(1).strip()

            # Extract DV number
            dv_m = re.search(r'dv-(\d+)', text)
            dv_num = int(dv_m.group(1)) if dv_m else 0

            if status.startswith("on notice"):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_NOTICE
            elif status == "vacant ready":
                ws.cell(row=row_idx, column=col_idx).fill = FILL_GREEN
            elif status == "vacant not ready":
                if dv_num >= SLA_DAYS:
                    ws.cell(row=row_idx, column=col_idx).fill = FILL_PINK
                else:
                    ws.cell(row=row_idx, column=col_idx).fill = FILL_GREEN


# ============================================================
# EXCEL WRITER
# ============================================================

def write_excel_file(path, sheets: dict):
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    table_counter = 0

    for sheet_name, sheet_spec in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        if isinstance(sheet_spec, pd.DataFrame):
            entries = [{"df": sheet_spec, "start_col": 1, "table_name": None}]
        elif isinstance(sheet_spec, list):
            entries = sheet_spec
        else:
            raise ValueError(f"Unsupported sheet spec type for '{sheet_name}'")

        for entry in entries:
            df = entry["df"]
            start_col = entry.get("start_col", 1)
            start_row = entry.get("start_row", 1)
            table_name = entry.get("table_name")

            table_counter += 1
            if table_name is None:
                table_name = f"Table{table_counter}"

            title = entry.get("title")
            if title:
                end_col_title = start_col + len(df.columns) - 1
                merge_top = start_row - 2
                merge_bot = start_row - 1
                ws.merge_cells(
                    start_row=merge_top,
                    start_column=start_col,
                    end_row=merge_bot,
                    end_column=end_col_title,
                )
                title_cell = ws.cell(row=merge_top, column=start_col, value=title)
                title_cell.font = Font(bold=True, size=20)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                thick = Side(style="thick")
                title_border = Border(top=thick, bottom=thick, left=thick, right=thick)
                for r in range(merge_top, merge_bot + 1):
                    for c in range(start_col, end_col_title + 1):
                        ws.cell(row=r, column=c).border = title_border

            date_fmt = "MM/DD/YY"

            date_col_indices = set()
            for ci, col_name in enumerate(df.columns):
                if pd.api.types.is_datetime64_any_dtype(df[col_name]):
                    date_col_indices.add(ci)
            for col_idx, col_name in enumerate(df.columns, start=start_col):
                ws.cell(row=start_row, column=col_idx, value=col_name)

            for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
                for col_idx, value in enumerate(row, start=start_col):
                    ci = col_idx - start_col
                    if pd.isna(value):
                        value = None
                    elif ci in date_col_indices and value is not None:
                        value = pd.to_datetime(value, errors="coerce")
                        if pd.isna(value):
                            value = None
                        else:
                            value = value.to_pydatetime()
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if ci in date_col_indices and value is not None:
                        cell.number_format = date_fmt

            end_col = start_col + len(df.columns) - 1
            end_row = start_row + len(df)

            ref = (
                f"{get_column_letter(start_col)}{start_row}:"
                f"{get_column_letter(end_col)}{end_row}"
            )

            # --- Table style: TableStyleMedium15 with row stripes ---
            style = TableStyleInfo(
                name="TableStyleMedium15",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            tbl = Table(displayName=table_name, ref=ref)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)

            # --- Header font: white bold ---
            header_font = Font(bold=True, color="FFFFFF")
            center = Alignment(horizontal="center", vertical="center")
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.font = header_font
                cell.alignment = center

            # --- Center all data cells ---
            for row_idx in range(start_row + 1, end_row + 1):
                for col_idx in range(start_col, end_col + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center

            # --- Dashboard conditional formatting ---
            if sheet_name == "Dashboard":
                col_map = {}
                for ci, col_name in enumerate(df.columns):
                    col_map[col_name] = start_col + ci
                _apply_dashboard_formatting(ws, start_row, end_row, col_map)

            # --- Operations conditional formatting ---
            if sheet_name == "Operations":
                col_map = {}
                for ci, col_name in enumerate(df.columns):
                    col_map[col_name] = start_col + ci
                _apply_operations_formatting(ws, start_row, end_row, col_map)

            # --- Aging conditional formatting ---
            if sheet_name == "Aging":
                _apply_aging_formatting(ws, start_row, end_row, start_col, end_col)

        # --- Autofit columns ---
        table_cols = set()
        for entry in entries:
            sc = entry.get("start_col", 1)
            ec = sc + len(entry["df"].columns) - 1
            table_cols.update(range(sc, ec + 1))

        max_col_used = max(table_cols) if table_cols else 1
        for col_num in range(1, max_col_used + 1):
            col_letter = get_column_letter(col_num)
            if col_num not in table_cols:
                ws.column_dimensions[col_letter].width = 4.0
            else:
                max_len = 0
                for row in ws.iter_rows(min_col=col_num, max_col=col_num, values_only=True):
                    val = row[0]
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(path)
