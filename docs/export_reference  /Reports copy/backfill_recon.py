import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ----------------------------------------
# PATHS (relative to this script's location)
# ----------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DMRB_PATH = os.path.join(SCRIPT_DIR, "data", "DMRB_raw.xlsx")
RECON_PATH = os.path.join(SCRIPT_DIR, "output", "Final_Report.xlsx")

# ----------------------------------------
# LOAD DATA
# ----------------------------------------
recon = pd.read_excel(RECON_PATH, sheet_name="Reconciliation")
recon["Unit"] = recon["Unit"].str.strip()

dmrb = pd.read_excel(DMRB_PATH, sheet_name="DMRB ")
dmrb.columns = dmrb.columns.str.strip()
dmrb["Unit"] = dmrb["Unit"].str.strip()

# Build a lookup from DMRB keyed by Unit
dmrb_lookup = dmrb.set_index("Unit")

# ----------------------------------------
# COLUMN MAPPING: Reconciliation col -> DMRB col
# ----------------------------------------
FILL_MAP = {
    "Status":              "Status",
    "Available Date":      "Move_out",
    "Move-In Ready Date":  "Ready_Date",
}

# ----------------------------------------
# BACKFILL BLANKS
# ----------------------------------------
filled_counts = {col: 0 for col in FILL_MAP}

for idx, row in recon.iterrows():
    unit = row["Unit"]
    if unit not in dmrb_lookup.index:
        continue
    dmrb_row = dmrb_lookup.loc[unit]
    # If duplicate units in DMRB, take the first match
    if isinstance(dmrb_row, pd.DataFrame):
        dmrb_row = dmrb_row.iloc[0]

    for recon_col, dmrb_col in FILL_MAP.items():
        if pd.isna(row[recon_col]) or (isinstance(row[recon_col], str) and row[recon_col].strip() == ""):
            dmrb_val = dmrb_row[dmrb_col]
            if pd.notna(dmrb_val) and not (isinstance(dmrb_val, str) and dmrb_val.strip() == ""):
                recon.at[idx, recon_col] = dmrb_val
                filled_counts[recon_col] += 1

# ----------------------------------------
# WRITE BACK TO Final_Report.xlsx (Reconciliation sheet only)
# ----------------------------------------
wb = load_workbook(RECON_PATH)
ws = wb["Reconciliation"]

# Build column index map from header row
header = {cell.value: cell.column for cell in ws[1]}

for idx, row in recon.iterrows():
    excel_row = idx + 2  # header is row 1
    for recon_col in FILL_MAP:
        col_num = header[recon_col]
        ws.cell(row=excel_row, column=col_num, value=row[recon_col])

# ----------------------------------------
# SHEET 2 – SPLIT VIEW (No Move In / Has Move In)
# ----------------------------------------
cols = ["phase", "Unit", "Status", "Available Date", "Move-In Ready Date",
        "Move In Date", "MO / Confirm"]

recon["Available Date"] = pd.to_datetime(recon["Available Date"], errors="coerce")
recon["Move-In Ready Date"] = pd.to_datetime(recon["Move-In Ready Date"], errors="coerce")
recon["Move In Date"] = pd.to_datetime(recon["Move In Date"], errors="coerce")

no_move_in = recon[recon["Move In Date"].isna()][cols].copy()
no_move_in = no_move_in.sort_values("Available Date").reset_index(drop=True)

has_move_in = recon[recon["Move In Date"].notna()][cols].copy()
has_move_in = has_move_in.sort_values(
    ["Move In Date", "Available Date"], ascending=[True, True]
).reset_index(drop=True)

# Remove old sheet if it exists
if "Split View" in wb.sheetnames:
    del wb["Split View"]

ws2 = wb.create_sheet("Split View")

# --- Write table 1 (has move in) header + data ---
for col_idx, col in enumerate(cols, start=1):
    ws2.cell(row=1, column=col_idx, value=col)
for r_idx, row in has_move_in.iterrows():
    for col_idx, col in enumerate(cols, start=1):
        ws2.cell(row=r_idx + 2, column=col_idx, value=row[col])

t1_last_row = len(has_move_in) + 1

# --- Spacer row ---
gap_row = t1_last_row + 1  # blank row

# --- Write table 2 (no move in) header + data ---
t2_header_row = gap_row + 1
for col_idx, col in enumerate(cols, start=1):
    ws2.cell(row=t2_header_row, column=col_idx, value=col)
for r_idx, row in no_move_in.iterrows():
    for col_idx, col in enumerate(cols, start=1):
        ws2.cell(row=t2_header_row + r_idx + 1, column=col_idx, value=row[col])

t2_last_row = t2_header_row + len(no_move_in)

# --- Excel tables ---
num_cols = len(cols)
last_col_letter = get_column_letter(num_cols)

t1 = Table(displayName="HasMoveIn", ref=f"A1:{last_col_letter}{t1_last_row}")
t1.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)
ws2.add_table(t1)

t2 = Table(displayName="NoMoveIn",
           ref=f"A{t2_header_row}:{last_col_letter}{t2_last_row}")
t2.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium15", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)
ws2.add_table(t2)

# --- Formatting ---
center = Alignment(horizontal="center", vertical="center")
header_font = Font(color="FFFFFF", bold=True)

all_data = pd.concat([no_move_in, has_move_in], ignore_index=True)
for i, col in enumerate(cols, start=1):
    max_len = max(all_data[col].fillna("").astype(str).str.len().max(), len(col)) + 2
    ws2.column_dimensions[get_column_letter(i)].width = max_len

for row in ws2.iter_rows(min_row=1, max_row=t2_last_row, min_col=1, max_col=num_cols):
    for cell in row:
        cell.alignment = center

for col_idx, col in enumerate(cols, start=1):
    if "Date" in col:
        for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx,
                                  max_row=t2_last_row):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "MM/DD/YY"

for cell in ws2[1]:
    cell.font = header_font
for cell in ws2[t2_header_row]:
    cell.font = header_font

wb.save(RECON_PATH)

# ----------------------------------------
# SUMMARY
# ----------------------------------------
print("Backfill complete!")
for col, count in filled_counts.items():
    print(f"  {col}: {count} blanks filled from DMRB")
print(f"\nSplit View: {len(no_move_in)} without Move In, {len(has_move_in)} with Move In")
print(f"\nUpdated: {RECON_PATH}")
