"""
Move-In Records Ledger

Captures every unit with a Move_in date from DMRB_raw.
Holds the record. Purges anything 30+ days past its Move_in date.
Output: two columns — Unit and Move-In Date (MM/DD/YYYY).
Formatted as an Excel table matching DMRB_REPORTS style.
"""

import pandas as pd
from pathlib import Path
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

PURGE_DAYS = 30


def update_movein_records(df_enriched: pd.DataFrame, output_path: Path):
    today = date.today()

    # --------------------------------------------------
    # Load existing ledger
    # --------------------------------------------------
    if output_path.exists():
        try:
            existing = pd.read_excel(output_path, sheet_name="Records")
            existing["Move-In Date"] = pd.to_datetime(existing["Move-In Date"], errors="coerce")
        except Exception:
            existing = pd.DataFrame(columns=["Unit", "Move-In Date"])
    else:
        existing = pd.DataFrame(columns=["Unit", "Move-In Date"])

    # --------------------------------------------------
    # Extract current units with Move_in dates
    # --------------------------------------------------
    df = df_enriched.copy()
    df["Move_in"] = pd.to_datetime(df["Move_in"], errors="coerce")

    has_movein = df[df["Move_in"].notna()][["Unit", "Move_in"]].copy()
    has_movein = has_movein.rename(columns={"Move_in": "Move-In Date"})

    if has_movein.empty and existing.empty:
        return None

    # --------------------------------------------------
    # Deduplicate: only add records not already in ledger
    # Key = Unit + Move-In Date
    # --------------------------------------------------
    if not existing.empty and not has_movein.empty:
        existing["_key"] = (
            existing["Unit"].astype(str) + "|" +
            existing["Move-In Date"].dt.strftime("%Y-%m-%d").fillna("")
        )
        has_movein["_key"] = (
            has_movein["Unit"].astype(str) + "|" +
            has_movein["Move-In Date"].dt.strftime("%Y-%m-%d").fillna("")
        )
        new_only = has_movein[~has_movein["_key"].isin(existing["_key"])]
        combined = pd.concat(
            [existing.drop(columns="_key"), new_only.drop(columns="_key")],
            ignore_index=True,
        )
    elif existing.empty:
        combined = has_movein.copy()
    else:
        combined = existing.copy()

    # --------------------------------------------------
    # Purge: remove records where Move-In Date is 30+ days ago
    # --------------------------------------------------
    combined["Move-In Date"] = pd.to_datetime(combined["Move-In Date"], errors="coerce")
    cutoff = pd.Timestamp(today - timedelta(days=PURGE_DAYS))
    combined = combined[
        combined["Move-In Date"].isna() | (combined["Move-In Date"] > cutoff)
    ]

    combined = combined.sort_values("Move-In Date", ascending=True, na_position="last")
    combined = combined.reset_index(drop=True)

    # --------------------------------------------------
    # Write Excel with DMRB_REPORTS table style
    # --------------------------------------------------
    _write_ledger(combined, output_path)

    return combined


def _write_ledger(df: pd.DataFrame, path: Path):
    path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Records"

    start_col = 2
    start_row = 5

    # Title
    end_col_title = start_col + len(df.columns) - 1
    merge_top = start_row - 2
    merge_bot = start_row - 1
    ws.merge_cells(
        start_row=merge_top,
        start_column=start_col,
        end_row=merge_bot,
        end_column=end_col_title,
    )
    title_cell = ws.cell(row=merge_top, column=start_col, value="Move-In Records")
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=col_name)

    # Data
    date_fmt = "MM/DD/YYYY"
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Move-In Date column (second column)
            if col_idx == start_col + 1 and value is not None:
                cell.number_format = date_fmt

    end_col = start_col + len(df.columns) - 1
    end_row = start_row + len(df)

    # Table
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    style = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tbl = Table(displayName="MoveInRecords", ref=ref)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = header_font
        cell.alignment = center

    # Center all data cells
    for row_idx in range(start_row + 1, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center

    # Autofit
    for col_num in range(1, end_col + 2):
        col_letter = get_column_letter(col_num)
        if col_num < start_col:
            ws.column_dimensions[col_letter].width = 4.0
        else:
            max_len = 0
            for row in ws.iter_rows(min_col=col_num, max_col=col_num, values_only=True):
                if row[0] is not None:
                    max_len = max(max_len, len(str(row[0])))
            ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(str(path))
