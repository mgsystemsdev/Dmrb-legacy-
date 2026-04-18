import os
import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MOVEIN_PATH = os.path.join(SCRIPT_DIR, "output", "MoveIn_Records.xlsx")
PURGE_DAYS = 30


# ----------------------------------------
# SHEET 1 – AVAILABLE UNITS
# ----------------------------------------
def process_available_units(input_path, writer):

    df = pd.read_csv(input_path, skiprows=5)

    df = df[["Unit", "Status", "Available Date", "Move-In Ready Date"]]

    df["Unit"] = df["Unit"].str.strip()

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )

    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[df["phase"].isin([5, 7, 8])]

    cols = ["phase"] + [c for c in df.columns if c != "phase"]
    df = df[cols]

    df["Available Date"] = pd.to_datetime(df["Available Date"], errors="coerce")
    df["Move-In Ready Date"] = pd.to_datetime(df["Move-In Ready Date"], errors="coerce")

    df = df.sort_values("Available Date").reset_index(drop=True)

    df.to_excel(writer, sheet_name="Available Units", index=False)

    style_sheet(writer, df, "Available Units", "AvailableUnits")

    return df


# ----------------------------------------
# SHEET 2 – MOVE INS
# ----------------------------------------
def process_move_ins(input_path, writer):

    df = pd.read_csv(input_path, skiprows=5, index_col=False)

    df = df[["Unit", "Move In Date"]]

    df["Unit"] = "Unit " + df["Unit"].str.strip()

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )

    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[df["phase"].isin([5, 7, 8])]

    cols = ["phase"] + [c for c in df.columns if c != "phase"]
    df = df[cols]

    df["Move In Date"] = pd.to_datetime(df["Move In Date"], errors="coerce")

    df = df.sort_values("Move In Date").reset_index(drop=True)

    df.to_excel(writer, sheet_name="Move Ins", index=False)

    style_sheet(writer, df, "Move Ins", "MoveIns")

    return df


# ----------------------------------------
# SHEET 3 – MOVE OUTS
# ----------------------------------------
def process_move_outs(input_path, writer):

    df = pd.read_csv(input_path, skiprows=6, index_col=False)

    df = df[["Unit", "Move-Out Date"]]

    df["Unit"] = "Unit " + df["Unit"].str.strip()

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )

    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[df["phase"].isin([5, 7, 8])]

    cols = ["phase"] + [c for c in df.columns if c != "phase"]
    df = df[cols]

    df["Move-Out Date"] = pd.to_datetime(df["Move-Out Date"], errors="coerce")

    df = df.sort_values("Move-Out Date").reset_index(drop=True)

    df.to_excel(writer, sheet_name="Move Outs", index=False)

    style_sheet(writer, df, "Move Outs", "MoveOuts")

    return df


# ----------------------------------------
# SHEET 4 – PENDING FAS
# ----------------------------------------
def process_pending_fas(input_path, writer, prev_completed):

    df = pd.read_csv(input_path, skiprows=4)

    df = df[["Unit Number", "MO / Cancel Date", "Lease End"]]

    df = df.rename(columns={"Unit Number": "Unit"})

    df["Unit"] = "Unit " + df["Unit"].str.strip()

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )

    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[df["phase"].isin([5, 7, 8])]

    cols = ["phase"] + [c for c in df.columns if c != "phase"]
    df = df[cols]

    df["MO / Cancel Date"] = pd.to_datetime(df["MO / Cancel Date"], errors="coerce")
    df["Lease End"] = pd.to_datetime(df["Lease End"], errors="coerce")

    # Restore user-typed "Completed" values for rows that still exist
    df["Completed"] = df["Unit"].map(prev_completed).fillna("")

    df = df.sort_values("MO / Cancel Date").reset_index(drop=True)

    df.to_excel(writer, sheet_name="Pending FAS", index=False)

    style_sheet(writer, df, "Pending FAS", "PendingFAS")

    return df


# ----------------------------------------
# SHEET 5 – MOVE ACTIVITY (INS + OUTS)
# ----------------------------------------
def process_move_activity(move_ins_df, move_outs_df, writer):

    df = pd.merge(
        move_ins_df[["Unit", "Move In Date"]],
        move_outs_df[["Unit", "Move-Out Date"]],
        on="Unit",
        how="outer"
    )

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )
    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[["phase", "Unit", "Move-Out Date", "Move In Date"]]

    df = df.sort_values("Unit").reset_index(drop=True)

    df.to_excel(writer, sheet_name="Move Activity", index=False)

    style_sheet(writer, df, "Move Activity", "MoveActivity")


# ----------------------------------------
# SHEET 6 – RECONCILIATION
# ----------------------------------------
def process_reconciliation(avail_df, move_ins_df, move_outs_df, fas_df, writer):

    avail = avail_df[["Unit", "Status", "Available Date", "Move-In Ready Date"]].copy()
    ins = move_ins_df[["Unit", "Move In Date"]].copy()
    fas = fas_df[["Unit", "MO / Cancel Date"]].copy()

    df = pd.merge(avail, ins, on="Unit", how="outer")
    df = pd.merge(df, fas, on="Unit", how="outer")

    confirmed = df["MO / Cancel Date"].notna()
    df.loc[confirmed, "Available Date"] = df.loc[confirmed, "MO / Cancel Date"]
    df.loc[confirmed, "Status"] = "Vacant Not Ready"
    df["MO / Confirm"] = confirmed.map({True: "Yes", False: ""})
    df = df.drop(columns=["MO / Cancel Date"])

    df["phase"] = (
        df["Unit"]
        .str.replace("Unit ", "", regex=False)
        .str.split("-")
        .str[0]
    )
    df["phase"] = pd.to_numeric(df["phase"], errors="coerce")

    df = df[["phase", "Unit", "Status", "Available Date", "Move-In Ready Date",
             "Move In Date", "MO / Confirm"]]

    df = df.sort_values(
        ["Move In Date", "Available Date"],
        ascending=[True, True]
    ).reset_index(drop=True)

    df.to_excel(writer, sheet_name="Reconciliation", index=False)

    style_sheet(writer, df, "Reconciliation", "Reconciliation")


# ----------------------------------------
# SHARED STYLING ENGINE
# ----------------------------------------
def style_sheet(writer, df, sheet_name, table_name):

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Autofit columns
    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].fillna("").astype(str).str.len().max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = max_len

    # Format date columns automatically
    for col_idx, col in enumerate(df.columns, start=1):
        if "Date" in col or "Lease End" in col:
            for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "MM/DD/YY"

    # Add Excel Table
    last_row = len(df) + 1
    last_col = get_column_letter(len(df.columns))

    table = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")

    style = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Center alignment
    center = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=last_row,
        min_col=1,
        max_col=len(df.columns)
    ):
        for cell in row:
            cell.alignment = center

    # Header styling
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.font = header_font


# ----------------------------------------
# MOVE-IN RECORDS LEDGER
# ----------------------------------------
def update_movein_records(move_ins_df):
    today = date.today()

    # Load existing ledger
    if os.path.exists(MOVEIN_PATH):
        try:
            existing = pd.read_excel(MOVEIN_PATH, sheet_name="Records")
            existing["Move-In Date"] = pd.to_datetime(existing["Move-In Date"], errors="coerce")
        except Exception:
            existing = pd.DataFrame(columns=["Unit", "Move-In Date"])
    else:
        existing = pd.DataFrame(columns=["Unit", "Move-In Date"])

    # Prepare current move-ins
    has_movein = move_ins_df[move_ins_df["Move In Date"].notna()][["Unit", "Move In Date"]].copy()
    has_movein = has_movein.rename(columns={"Move In Date": "Move-In Date"})

    if has_movein.empty and existing.empty:
        return

    # Deduplicate
    if not existing.empty and not has_movein.empty:
        existing["_key"] = (
            existing["Unit"].astype(str) + "|" +
            existing["Move-In Date"].dt.strftime("%Y-%m-%d").fillna("")
        )
        has_movein["_key"] = (
            has_movein["Unit"].astype(str) + "|" +
            has_movein["Move-In Date"].dt.strftime("%Y-%m-%d").fillna("")
        )
        new_only = has_movein[~has_movein["_key"].isin(existing["_key"])]
        combined = pd.concat(
            [existing.drop(columns="_key"), new_only.drop(columns="_key")],
            ignore_index=True,
        )
    elif existing.empty:
        combined = has_movein.copy()
    else:
        combined = existing.copy()

    # Purge 30+ days ago
    combined["Move-In Date"] = pd.to_datetime(combined["Move-In Date"], errors="coerce")
    cutoff = pd.Timestamp(today - timedelta(days=PURGE_DAYS))
    combined = combined[
        combined["Move-In Date"].isna() | (combined["Move-In Date"] > cutoff)
    ]
    combined = combined.sort_values("Move-In Date", ascending=True, na_position="last")
    combined = combined.reset_index(drop=True)

    # Write styled Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"

    start_col = 2
    start_row = 5

    # Title
    end_col_title = start_col + len(combined.columns) - 1
    ws.merge_cells(
        start_row=start_row - 2, start_column=start_col,
        end_row=start_row - 1, end_column=end_col_title,
    )
    title_cell = ws.cell(row=start_row - 2, column=start_col, value="Move-In Records")
    title_cell.font = Font(bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    for col_idx, col_name in enumerate(combined.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=col_name)

    # Data
    for row_idx, row in enumerate(combined.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == start_col + 1 and value is not None:
                cell.number_format = "MM/DD/YYYY"

    end_col = start_col + len(combined.columns) - 1
    end_row = start_row + len(combined)

    # Table
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    tbl = Table(displayName="MoveInRecords", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = header_font
        cell.alignment = center
    for row_idx in range(start_row + 1, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center

    # Autofit
    for col_num in range(1, end_col + 2):
        col_letter = get_column_letter(col_num)
        if col_num < start_col:
            ws.column_dimensions[col_letter].width = 4.0
        else:
            max_len = 0
            for row in ws.iter_rows(min_col=col_num, max_col=col_num, values_only=True):
                if row[0] is not None:
                    max_len = max(max_len, len(str(row[0])))
            ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(MOVEIN_PATH)
    print(f"📋 Move-In Records: {len(combined)} records → {MOVEIN_PATH}")


# ----------------------------------------
# MASTER FILE
# ----------------------------------------
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "output", "Final_Report.xlsx")

# Read previous "Completed" values BEFORE the writer overwrites the file
prev_completed = {}
if os.path.exists(OUTPUT_FILE):
    try:
        prev = pd.read_excel(OUTPUT_FILE, sheet_name="Pending FAS")
        if "Completed" in prev.columns:
            prev["Unit"] = prev["Unit"].str.strip()
            for _, r in prev.iterrows():
                val = r["Completed"]
                if pd.notna(val) and str(val).strip() != "":
                    prev_completed[r["Unit"]] = val
    except Exception:
        pass

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    avail_df = process_available_units(os.path.join(SCRIPT_DIR, "data", "Available Units.csv"), writer)
    move_ins_df = process_move_ins(os.path.join(SCRIPT_DIR, "data", "Pending Move Ins.csv"), writer)
    move_outs_df = process_move_outs(os.path.join(SCRIPT_DIR, "data", "Move-Outs.csv"), writer)
    fas_df = process_pending_fas(os.path.join(SCRIPT_DIR, "data", "PendingFAS-.csv"), writer, prev_completed)
    process_move_activity(move_ins_df, move_outs_df, writer)
    process_reconciliation(avail_df, move_ins_df, move_outs_df, fas_df, writer)

    writer.book.move_sheet("Reconciliation", offset=-5)

update_movein_records(move_ins_df)

