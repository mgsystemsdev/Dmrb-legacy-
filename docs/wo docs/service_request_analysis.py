import os
import pandas as pd

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MOVEIN_PATH = os.path.join(SCRIPT_DIR, "Data", "MoveIn_Records.xlsx")
SR_PATH = os.path.join(SCRIPT_DIR, "Data", "Active Service Request.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "Output", "Service_Request_Report.xlsx")


# ----------------------------------------
# SHARED STYLING ENGINE
# ----------------------------------------
def style_sheet(writer, df, sheet_name, table_name):
    worksheet = writer.sheets[sheet_name]

    if len(df) == 0:
        return

    # Autofit columns
    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].fillna("").astype(str).str.len().max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = min(max_len, 40)

    # Format date columns
    for col_idx, col in enumerate(df.columns, start=1):
        if "Date" in col or "date" in col.lower():
            for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "MM/DD/YY"

    # Add Excel Table
    last_row = len(df) + 1
    last_col = get_column_letter(len(df.columns))
    table = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    # Center alignment
    center = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(
        min_row=1, max_row=last_row, min_col=1, max_col=len(df.columns)
    ):
        for cell in row:
            cell.alignment = center

    # Header styling
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.font = header_font


# ----------------------------------------
# STEP 1 – LOAD MOVE-IN RECORDS
# ----------------------------------------
def load_movein_records(path):
    df = pd.read_excel(path, sheet_name="Records", header=None, skiprows=4)
    df = df.dropna(axis=1, how="all")
    df.columns = ["Unit", "Move-In Date"]
    df["Unit"] = df["Unit"].str.strip()
    df["Move-In Date"] = pd.to_datetime(df["Move-In Date"], errors="coerce")
    df = df[df["Unit"].notna() & df["Move-In Date"].notna()]
    # Strip "Unit " prefix to match SR Location format (e.g. "5-18-0106")
    df["Unit"] = df["Unit"].str.replace("Unit ", "", regex=False)
    return df[["Unit", "Move-In Date"]]


# ----------------------------------------
# STEP 2 – LOAD SERVICE REQUESTS
# ----------------------------------------
def load_service_requests(path):
    df = pd.read_excel(path)
    df = df[["Location", "Created date", "Due date", "Service Category",
             "Issue", "Assigned to", "Status"]]
    df["Location"] = df["Location"].str.strip()

    # Extract phase from Location (first segment before "-")
    df["Phase"] = df["Location"].str.split("-").str[0].str.strip().str.upper()

    # Place Phase before Location
    cols = ["Phase"] + [c for c in df.columns if c != "Phase"]
    df = df[cols]

    df["Created date"] = pd.to_datetime(df["Created date"], format="%m/%d/%Y", errors="coerce")
    df["Due date"] = pd.to_datetime(df["Due date"], format="%m/%d/%Y", errors="coerce")
    return df


# ----------------------------------------
# STEP 3 – MERGE & COMPUTE
# ----------------------------------------
def merge_and_compute(sr_df, units_df):
    df = sr_df.merge(
        units_df[["Unit", "Move-In Date"]],
        left_on="Location",
        right_on="Unit",
        how="left",
    )
    df = df.drop(columns=["Unit"])

    # Days Since Move-In
    df["Days Since Move-In"] = (df["Created date"] - df["Move-In Date"]).dt.days
    df.loc[df["Move-In Date"].isna(), "Days Since Move-In"] = None

    # Make Ready Request – flag if SR was created within 14 days of move-in
    df["Make Ready Request"] = ""
    mask = df["Days Since Move-In"].between(-14, 14)
    df.loc[mask, "Make Ready Request"] = "Yes"

    return df


# ----------------------------------------
# SHEET 1 – ALL SERVICE REQUESTS (merged)
# ----------------------------------------
def write_all_requests(df, writer):
    out = df.copy().reset_index(drop=True)
    out.to_excel(writer, sheet_name="All Requests", index=False)
    style_sheet(writer, out, "All Requests", "AllRequests")
    return out


# ----------------------------------------
# SHEET 2 – MAKE READY ONLY
# ----------------------------------------
def write_make_ready(df, writer):
    mr = df[df["Make Ready Request"] == "Yes"].copy().reset_index(drop=True)
    mr.to_excel(writer, sheet_name="Make Ready", index=False)
    style_sheet(writer, mr, "Make Ready", "MakeReady")

    # Highlight Make Ready Request column
    ws = writer.sheets["Make Ready"]
    if len(mr) > 0:
        col_idx = list(mr.columns).index("Make Ready Request") + 1
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=len(mr) + 1,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.fill = green_fill

    return mr


# ----------------------------------------
# MASTER FILE
# ----------------------------------------
movein_df = load_movein_records(MOVEIN_PATH)
sr_df = load_service_requests(SR_PATH)
merged_df = merge_and_compute(sr_df, movein_df)

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    all_df = write_all_requests(merged_df, writer)
    mr_df = write_make_ready(merged_df, writer)

# Summary
total = len(merged_df)
matched = merged_df["Move-In Date"].notna().sum()
make_ready = len(mr_df)

print(f"Total service requests:   {total}")
print(f"Matched to a unit:        {matched}")
print(f"Unmatched (common areas):  {total - matched}")
print(f"Make Ready requests:       {make_ready}")
print(f"\nOutput saved to: {OUTPUT_PATH}")
